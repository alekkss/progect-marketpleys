"""
Основной класс синхронизации данных между маркетплейсами
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from typing import Dict, Tuple, Optional, List
from pathlib import Path
import sys

from config.config import FILE_CONFIGS, is_excluded_column
from utils.logger_config import setup_logger
from .constants import ARTICLE_COLUMNS, DIMENSIONS_MAPPING
from .converters import ValueConverter
from .dimensions import DimensionsSynchronizer
from .alignment import ArticleAligner
from .validation import ValidationChain

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

logger = setup_logger('data_sync')


class DataSynchronizer:
    """Класс для синхронизации данных между тремя маркетплейсами"""
    
    def __init__(self, comparison_result: Dict, ai_comparator=None):
        """
        Args:
            comparison_result: результат сравнения столбцов из AIComparator
            ai_comparator: экземпляр AIComparator для AI-валидации
        """
        self.comparison_result = comparison_result
        self.article_columns = ARTICLE_COLUMNS
        self.changes_log = {
            'wildberries': [],
            'ozon': [],
            'yandex': []
        }
        self.original_file_paths = {}
        self.ai_comparator = ai_comparator
        
        # Валидация
        self.validation_chain = ValidationChain(ai_comparator)
        self.column_validations = {}  # {marketplace: {column_name: [allowed_values]}}
        self.original_column_names = {}
        
        logger.info("Инициализация DataSynchronizer")
        logger.debug(f"AI comparator передан: {ai_comparator is not None}")
    
    @property
    def ai_validation_log(self):
        """Getter для логов AI-валидации"""
        return self.validation_chain.ai_validation_log
    
    def synchronize_data(
        self,
        file_paths: Dict[str, str],
        output_paths: Dict[str, str] = None,
        report_path: str = None
    ) -> Tuple[Dict[str, pd.DataFrame], Dict]:
        """
        Основной метод синхронизации данных
        
        Args:
            file_paths: пути к файлам маркетплейсов
            output_paths: пути для сохранения результатов
            report_path: путь к отчёту
            
        Returns:
            Кортеж (синхронизированные DataFrame, лог изменений)
        """
        logger.info("="*60)
        logger.info("СИНХРОНИЗАЦИЯ ДАННЫХ МЕЖДУ МАРКЕТПЛЕЙСАМИ")
        logger.info("="*60)
        
        # 1. Загружаем данные из всех трех файлов
        dfs = self._load_all_dataframes(file_paths)
        
        # 2. Выравниваем артикулы
        logger.info("\n[*] Выравнивание артикулов...")
        dfs = ArticleAligner.align_articles(dfs)
        
        # 3. Синхронизация композитных габаритов
        logger.info("\n[*] Синхронизация габаритов...")
        dimensions_synced = DimensionsSynchronizer.sync_dimensions(dfs)
        
        # 4. Синхронизируем данные по схеме
        synced_dfs = self._sync_all_matches(dfs)
        
        # 5. Постобработка габаритов WB (мм → см если из Ozon)
        logger.info("\n[*] Постобработка габаритов WB...")
        converted_count = self._postprocess_wb_dimensions(synced_dfs)
        if converted_count > 0:
            logger.info(f"✅ Сконвертировано {converted_count} значений габаритов (мм → см)")
        
        # 6. Сохраняем результаты
        if output_paths:
            self._save_results(synced_dfs, output_paths)
        
        logger.info("✅ Синхронизация завершена!")
        return synced_dfs, self.changes_log
    
    def _load_all_dataframes(self, file_paths: Dict[str, str]) -> Dict[str, pd.DataFrame]:
        """Загружает данные через openpyxl для сохранения форматов"""
        logger.info("📂 Загружаю данные из файлов...")
        dfs = {}
        self.original_column_names = {}
        
        for marketplace, file_path in file_paths.items():
            self.original_file_paths[marketplace] = file_path
            config = FILE_CONFIGS[marketplace]
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb[config['sheet_name']]
            
            # Загружаем validation правила
            self._load_column_validations(ws, marketplace, config)
            
            # Читаем данные
            data = []
            headers = []
            
            # Читаем заголовки
            for cell in ws[config['header_row']]:
                headers.append(cell.value if cell.value else '')
            
            # Обработка дубликатов столбцов
            headers = self._handle_duplicate_columns(headers, marketplace)
            
            # Читаем данные
            data_start = config.get('data_start_row', config['header_row'] + 1)
            for row in ws.iter_rows(min_row=data_start, values_only=True):
                data.append(row)
            
            df = pd.DataFrame(data, columns=headers)
            dfs[marketplace] = df
            wb.close()
            
            logger.info(f"✅ {config['display_name']}: загружено {len(df)} товаров")
        
        return dfs
    
    def _handle_duplicate_columns(self, headers: List[str], marketplace: str) -> List[str]:
        """Обрабатывает дубликаты столбцов, добавляя суффиксы"""
        original_headers = headers.copy()
        seen = {}
        renamed_columns = {}
        
        for i, col in enumerate(headers):
            if col in seen:
                # Нашли дубликат - добавляем суффикс
                seen[col] += 1
                new_name = f"{col}{seen[col]}"
                logger.warning(f"⚠️ [{marketplace}] Дубликат столбца '{col}' переименован в '{new_name}'")
                headers[i] = new_name
                renamed_columns[new_name] = col
            else:
                seen[col] = 0
        
        # Сохраняем информацию для восстановления
        if renamed_columns:
            self.original_column_names[marketplace] = {
                'renamed': renamed_columns,
                'all_headers': original_headers
            }
        
        return headers
    
    def _load_column_validations(self, ws, marketplace: str, config: Dict):
        """Загружает информацию о validation для каждого столбца"""
        if marketplace not in self.column_validations:
            self.column_validations[marketplace] = {}
        
        header_row = config['header_row']
        
        # Создаем маппинг: номер колонки -> название
        col_idx_to_name = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                col_name = str(cell.value).strip()
                col_idx_to_name[col_idx] = col_name
        
        logger.info(f"📋 [{marketplace}] Найдено {len(col_idx_to_name)} столбцов")
        
        # Получаем именованные диапазоны
        workbook = ws.parent
        named_ranges = {}
        try:
            for name_obj in workbook.defined_names.values():
                try:
                    if name_obj.value:
                        named_ranges[name_obj.name] = name_obj.value
                except Exception as e:
                    logger.debug(f"[{marketplace}] Пропущен именованный диапазон: {e}")
            
            logger.info(f"[{marketplace}] Найдено {len(named_ranges)} именованных диапазонов")
        except Exception as e:
            logger.error(f"[{marketplace}] Ошибка получения именованных диапазонов: {e}")
        
        # Проходим по всем validation правилам
        validation_count = 0
        for dv_index, dv in enumerate(ws.data_validations.dataValidation, start=1):
            if dv.type != "list" or dv.sqref is None:
                continue
            
            # Извлекаем значения из validation
            allowed_values = self._extract_validation_values(dv, ws, workbook, named_ranges, marketplace, dv_index)
            
            if not allowed_values:
                continue
            
            # Определяем какие столбцы затронуты этим validation
            ranges = str(dv.sqref).split()
            for range_str in ranges:
                try:
                    if ':' in range_str:
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        
                        # Применяем validation ко всем колонкам в диапазоне
                        for col_idx in range(min_col, max_col + 1):
                            if col_idx in col_idx_to_name:
                                col_name = col_idx_to_name[col_idx]
                                self.column_validations[marketplace][col_name] = allowed_values
                                validation_count += 1
                                logger.info(f"✅ [{marketplace}] Validation для '{col_name}': {len(allowed_values)} значений")
                except Exception as e:
                    logger.error(f"[{marketplace}] Ошибка обработки range_str '{range_str}': {e}")
        
        logger.info(f"📊 [{marketplace}] Итого загружено validation для {validation_count} столбцов")
    
    def _extract_validation_values(
        self, 
        dv, 
        ws, 
        workbook, 
        named_ranges: Dict, 
        marketplace: str, 
        dv_index: int
    ) -> List[str]:
        """Извлекает значения из правила validation"""
        allowed_values = []
        
        if not dv.formula1:
            return allowed_values
        
        formula = dv.formula1
        
        # Список задан прямо: "Красный,Синий,Зеленый"
        if formula.startswith('"') and formula.endswith('"'):
            allowed_values = [v.strip() for v in formula.strip('"').split(',')]
            logger.debug(f"[{marketplace}] DV #{dv_index}: Прямой список, {len(allowed_values)} значений")
        
        # Именованный диапазон
        elif formula in named_ranges:
            try:
                range_formula = named_ranges[formula]
                clean_formula = range_formula.replace('$', '')
                
                if '!' in clean_formula:
                    sheet_name, range_ref = clean_formula.split('!', 1)
                    sheet_name = sheet_name.strip("'")
                    target_ws = workbook[sheet_name]
                else:
                    range_ref = clean_formula
                    target_ws = ws
                
                # Извлекаем значения
                for row in target_ws[range_ref]:
                    for cell in row:
                        if cell.value is not None:
                            allowed_values.append(str(cell.value).strip())
                
                logger.info(f"✅ [{marketplace}] DV #{dv_index}: Извлечено {len(allowed_values)} значений из '{formula}'")
            except Exception as e:
                logger.error(f"[{marketplace}] DV #{dv_index}: Ошибка обработки именованного диапазона '{formula}': {e}")
        
        # Обычный диапазон
        elif ':' in formula:
            try:
                clean_formula = formula.replace('$', '')
                if '!' in clean_formula:
                    sheet_name, range_ref = clean_formula.split('!')
                    target_ws = workbook[sheet_name]
                else:
                    range_ref = clean_formula
                    target_ws = ws
                
                for row in target_ws[range_ref]:
                    for cell in row:
                        if cell.value is not None:
                            allowed_values.append(str(cell.value).strip())
            except Exception as e:
                logger.error(f"[{marketplace}] DV #{dv_index}: Ошибка извлечения validation: {e}")
        
        return allowed_values
    
    def _sync_all_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Синхронизирует все совпадающие столбцы"""
        # Создаем копии для работы
        synced_dfs = {
            'wildberries': dfs['wildberries'].copy(),
            'ozon': dfs['ozon'].copy(),
            'yandex': dfs['yandex'].copy()
        }
        
        # Синхронизируем совпадения всех трех маркетплейсов
        print("\n[*] Синхронизирую совпадения всех 3 маркетплейсов...")
        synced_dfs = self._sync_three_way_matches(synced_dfs)
        
        # Синхронизируем совпадения между двумя маркетплейсами
        print("\n[*] Синхронизирую совпадения между парами маркетплейсов...")
        synced_dfs = self._sync_two_way_matches(synced_dfs)
        
        return synced_dfs
    
    def _sync_three_way_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Синхронизирует совпадения всех трех маркетплейсов"""
        matches = self.comparison_result.get('matches_all_three', [])
        
        if not matches:
            print(" Нет совпадений для синхронизации")
            return dfs
        
        total_filled = 0
        skipped_count = 0
        
        for match in matches:
            col_wb = match.get('column_1')
            col_ozon = match.get('column_2')
            col_yandex = match.get('column_3')
            
            if not all([col_wb, col_ozon, col_yandex]):
                continue
            
            # Пропускаем исключенные столбцы
            if (is_excluded_column(col_wb) or
                is_excluded_column(col_ozon) or
                is_excluded_column(col_yandex)):
                skipped_count += 1
                continue
            
            # Проверяем, что столбцы существуют
            if (col_wb not in dfs['wildberries'].columns or
                col_ozon not in dfs['ozon'].columns or
                col_yandex not in dfs['yandex'].columns):
                continue
            
            # Синхронизируем данные между тремя файлами
            filled = self._sync_three_columns(dfs, col_wb, col_ozon, col_yandex)
            
            if filled > 0:
                confidence = int(match.get('confidence', 0) * 100)
                print(f" ✓ Заполнено {filled} значений: '{col_wb}' ↔ '{col_ozon}' ↔ '{col_yandex}' ({confidence}%)")
                total_filled += filled
        
        if skipped_count > 0:
            print(f"[!] Пропущено {skipped_count} исключенных столбцов")
        
        print(f"[+] Всего заполнено {total_filled} пустых ячеек в совпадениях всех 3 маркетплейсов")
        return dfs
    
    def _sync_two_way_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Синхронизирует совпадения между парами маркетплейсов"""
        pairs = [
            ('matches_1_2', 'wildberries', 'ozon', 'column_1', 'column_2'),
            ('matches_1_3', 'wildberries', 'yandex', 'column_1', 'column_3'),
            ('matches_2_3', 'ozon', 'yandex', 'column_2', 'column_3')
        ]
        
        total_filled = 0
        skipped_count = 0
        
        for match_key, mp1, mp2, col_key1, col_key2 in pairs:
            matches = self.comparison_result.get(match_key, [])
            
            if not matches:
                continue
            
            for match in matches:
                col1 = match.get(col_key1)
                col2 = match.get(col_key2)
                
                if not all([col1, col2]):
                    continue
                
                # Пропускаем исключенные столбцы
                if is_excluded_column(col1) or is_excluded_column(col2):
                    skipped_count += 1
                    continue
                
                # Проверяем, что столбцы существуют
                if col1 not in dfs[mp1].columns or col2 not in dfs[mp2].columns:
                    continue
                
                # Синхронизируем данные между двумя файлами
                filled = self._sync_two_columns(dfs, mp1, mp2, col1, col2)
                
                if filled > 0:
                    confidence = int(match.get('confidence', 0) * 100)
                    print(f" ✓ Заполнено {filled} значений: {mp1}:'{col1}' ↔ {mp2}:'{col2}' ({confidence}%)")
                    total_filled += filled
        
        if skipped_count > 0:
            print(f"[!] Пропущено {skipped_count} исключенных столбцов")
        
        print(f"[+] Всего заполнено {total_filled} пустых ячеек в совпадениях между парами")
        return dfs
    
    def _sync_three_columns(
        self,
        dfs: Dict[str, pd.DataFrame],
        col_wb: str,
        col_ozon: str,
        col_yandex: str
    ) -> int:
        """
        Синхронизирует данные между тремя столбцами на основе артикулов
        
        Returns:
            Количество заполненных ячеек
        """
        filled_count = 0
        
        # Определяем единицы измерения для каждого столбца
        unit_wb = ValueConverter.detect_unit(col_wb)
        unit_ozon = ValueConverter.detect_unit(col_ozon)
        unit_yandex = ValueConverter.detect_unit(col_yandex)
        
        # Создаем словари для быстрого поиска по артикулу
        wb_data = self._create_article_map(dfs['wildberries'], ARTICLE_COLUMNS['wildberries'], col_wb)
        ozon_data = self._create_article_map(dfs['ozon'], ARTICLE_COLUMNS['ozon'], col_ozon)
        yandex_data = self._create_article_map(dfs['yandex'], ARTICLE_COLUMNS['yandex'], col_yandex)
        
        # Получаем все уникальные артикулы
        all_articles = set(wb_data.keys()) | set(ozon_data.keys()) | set(yandex_data.keys())
        
        for article in all_articles:
            if not article:
                continue
            
            # Получаем значения из всех трех источников
            values = {
                'wildberries': wb_data.get(article, {}).get('value'),
                'ozon': ozon_data.get(article, {}).get('value'),
                'yandex': yandex_data.get(article, {}).get('value')
            }
            
            # Находим непустое значение и его источник
            source_value, source_unit = self._find_source_value(values, unit_wb, unit_ozon, unit_yandex)
            
            if source_value is None:
                continue
            
            # Заполняем пустые значения в каждом маркетплейсе
            filled_count += self._fill_marketplace_value(
                dfs['wildberries'], article, col_wb, source_value, source_unit, unit_wb, 
                'wildberries', wb_data, values['wildberries']
            )
            
            filled_count += self._fill_marketplace_value(
                dfs['ozon'], article, col_ozon, source_value, source_unit, unit_ozon,
                'ozon', ozon_data, values['ozon']
            )
            
            # Для Яндекса проверяем композитные габариты
            if col_yandex == DIMENSIONS_MAPPING['yandex']['composite']:
                filled = self._fill_composite_dimensions(
                    dfs, article, col_yandex, source_unit, unit_wb, unit_ozon, yandex_data, values['yandex']
                )
                filled_count += filled
            else:
                filled_count += self._fill_marketplace_value(
                    dfs['yandex'], article, col_yandex, source_value, source_unit, unit_yandex,
                    'yandex', yandex_data, values['yandex']
                )
        
        return filled_count
    
    def _sync_two_columns(
        self,
        dfs: Dict[str, pd.DataFrame],
        mp1: str,
        mp2: str,
        col1: str,
        col2: str
    ) -> int:
        """Синхронизирует данные между двумя столбцами"""
        filled_count = 0
        
        unit1 = ValueConverter.detect_unit(col1)
        unit2 = ValueConverter.detect_unit(col2)
        
        data1 = self._create_article_map(dfs[mp1], ARTICLE_COLUMNS[mp1], col1)
        data2 = self._create_article_map(dfs[mp2], ARTICLE_COLUMNS[mp2], col2)
        
        all_articles = set(data1.keys()) | set(data2.keys())
        
        for article in all_articles:
            if not article:
                continue
            
            val1 = data1.get(article, {}).get('value')
            val2 = data2.get(article, {}).get('value')
            
            # Извлекаем скаляр если Series
            if isinstance(val1, pd.Series):
                val1 = val1.iloc[0] if not val1.empty else None
            if isinstance(val2, pd.Series):
                val2 = val2.iloc[0] if not val2.empty else None
            
            # Заполняем mp1 из mp2
            if (pd.isna(val1) or not str(val1).strip()) and pd.notna(val2) and str(val2).strip():
                filled_count += self._fill_marketplace_value(
                    dfs[mp1], article, col1, val2, unit2, unit1, mp1, data1, val1
                )
            
            # Заполняем mp2 из mp1
            if (pd.isna(val2) or not str(val2).strip()) and pd.notna(val1) and str(val1).strip():
                filled_count += self._fill_marketplace_value(
                    dfs[mp2], article, col2, val1, unit1, unit2, mp2, data2, val2
                )
        
        return filled_count
    
    def _find_source_value(self, values: Dict, unit_wb, unit_ozon, unit_yandex):
        """Находит непустое значение и его источник"""
        for marketplace, val in values.items():
            if isinstance(val, pd.Series):
                val = val.iloc[0] if not val.empty else None
            
            if pd.notna(val) and str(val).strip():
                if marketplace == 'wildberries':
                    return val, unit_wb
                elif marketplace == 'ozon':
                    return val, unit_ozon
                else:
                    return val, unit_yandex
        
        return None, None
    
    def _fill_marketplace_value(
        self, df: pd.DataFrame, article: str, col: str,
        source_value, source_unit, target_unit, marketplace: str,
        data_map: Dict, current_value
    ) -> int:
        """Заполняет значение в маркетплейсе с конвертацией и валидацией"""
        if isinstance(current_value, pd.Series):
            current_value = current_value.iloc[0] if not current_value.empty else None
        
        if pd.notna(current_value) and str(current_value).strip():
            return 0  # Уже заполнено
        
        if article not in data_map:
            return 0
        
        idx = data_map[article]['index']
        
        # Конвертируем значение
        converted_value = ValueConverter.convert_value(source_value, source_unit, target_unit)
        
        # Валидация
        allowed_values = self.column_validations.get(marketplace, {}).get(col)
        final_value = self.validation_chain.validate_multiple_values(
            converted_value, marketplace, col, allowed_values
        )
        
        # Решение о записи
        if final_value:
            value_to_set = final_value
        elif not allowed_values:
            value_to_set = converted_value
        else:
            logger.warning(f"⚠️ [{marketplace.upper()}] Пропущено '{converted_value}' для '{col}' (не прошло validation)")
            return 0
        
        # Записываем значение
        try:
            series = df[col]
            if isinstance(series, pd.DataFrame):
                series = series.iloc[:, 0]
            
            if pd.api.types.is_numeric_dtype(series.dtype):
                value_to_set = pd.to_numeric(value_to_set, errors='coerce')
            
            df.at[idx, col] = value_to_set
            self._log_change(marketplace, article, col, value_to_set)
            return 1
        except Exception as e:
            logger.error(f"Ошибка записи значения: {e}")
            return 0
    
    def _fill_composite_dimensions(
        self, dfs: Dict, article: str, col_yandex: str,
        source_unit, unit_wb, unit_ozon, yandex_data: Dict, current_value
    ) -> int:
        """Заполняет композитные габариты в Яндекс"""
        if isinstance(current_value, pd.Series):
            current_value = current_value.iloc[0] if not current_value.empty else None
        
        if pd.notna(current_value) and str(current_value).strip():
            return 0
        
        if article not in yandex_data:
            return 0
        
        idx = yandex_data[article]['index']
        composite = None
        
        # Из WB
        if source_unit == unit_wb:
            wb_row = dfs['wildberries'][
                dfs['wildberries'][ARTICLE_COLUMNS['wildberries']].astype(str).str.strip() == article
            ]
            if not wb_row.empty:
                wb_row = wb_row.iloc[0]
                length = wb_row.get(DIMENSIONS_MAPPING['wildberries']['length'])
                width = wb_row.get(DIMENSIONS_MAPPING['wildberries']['width'])
                height = wb_row.get(DIMENSIONS_MAPPING['wildberries']['height'])
                
                if all(pd.notna(v) for v in [length, width, height]):
                    composite = DimensionsSynchronizer.format_composite_dimensions(
                        float(length), float(width), float(height)
                    )
        
        # Из Ozon
        elif source_unit == unit_ozon:
            ozon_row = dfs['ozon'][
                dfs['ozon'][ARTICLE_COLUMNS['ozon']].astype(str).str.strip() == article
            ]
            if not ozon_row.empty:
                ozon_row = ozon_row.iloc[0]
                length_mm = ozon_row.get(DIMENSIONS_MAPPING['ozon']['length'])
                width_mm = ozon_row.get(DIMENSIONS_MAPPING['ozon']['width'])
                height_mm = ozon_row.get(DIMENSIONS_MAPPING['ozon']['height'])
                
                if all(pd.notna(v) for v in [length_mm, width_mm, height_mm]):
                    composite = DimensionsSynchronizer.format_composite_dimensions(
                        ValueConverter.mm_to_cm(float(length_mm)),
                        ValueConverter.mm_to_cm(float(width_mm)),
                        ValueConverter.mm_to_cm(float(height_mm))
                    )
        
        if composite:
            dfs['yandex'].at[idx, col_yandex] = composite
            self._log_change('yandex', article, col_yandex, composite)
            return 1
        
        return 0
    
    def _create_article_map(self, df: pd.DataFrame, article_col: str, value_col: str) -> Dict:
        """Создает маппинг артикул -> {index, value}"""
        article_map = {}
        
        for idx, row in df.iterrows():
            article = row.get(article_col)
            if pd.notna(article) and str(article).strip():
                article_str = str(article).strip()
                article_map[article_str] = {
                    'index': idx,
                    'value': row.get(value_col)
                }
        
        return article_map
    
    def _postprocess_wb_dimensions(self, dfs: Dict[str, pd.DataFrame]) -> int:
        """Постобработка габаритов WB: конвертация мм → см если значения из Ozon"""
        if 'wildberries' not in dfs:
            return 0
        
        converted_count = 0
        wb_map = DIMENSIONS_MAPPING['wildberries']
        df_wb = dfs['wildberries']
        
        for col in [wb_map['length'], wb_map['width'], wb_map['height']]:
            if col not in df_wb.columns:
                continue
            
            for idx, value in df_wb[col].items():
                if pd.notna(value):
                    try:
                        numeric_value = float(value)
                        # Если значение > 100, скорее всего это миллиметры
                        if numeric_value > 100:
                            df_wb.at[idx, col] = ValueConverter.mm_to_cm(numeric_value)
                            converted_count += 1
                    except (ValueError, TypeError):
                        pass
        
        return converted_count
    
    def _save_results(self, synced_dfs: Dict[str, pd.DataFrame], output_paths: Dict[str, str]):
        """Сохраняет результаты с сохранением форматов"""
        from utils.excel_writer import ExcelWriter
        writer = ExcelWriter()
        
        for marketplace, output_path in output_paths.items():
            if marketplace in synced_dfs:
                original_path = self.original_file_paths.get(marketplace)
                if original_path:
                    config = FILE_CONFIGS[marketplace]
                    writer.save_with_formatting(
                        synced_dfs[marketplace],
                        original_path,
                        output_path,
                        config['sheet_name'],
                        config['header_row']
                    )
                    logger.info(f"✅ Сохранен: {output_path}")
    
    def _log_change(self, marketplace: str, article: str, column: str, new_value, source_marketplace: str = None):
        """Логирует изменение"""
        self.changes_log[marketplace].append({
            'article': article,
            'column': column,
            'new_value': new_value,
            'source': source_marketplace
        })
    
    def _create_ai_log_sheet_in_report(self, report_path: str):
        """Создает лист с AI-логами в отчете"""
        if not self.ai_validation_log:
            return
        
        from openpyxl import load_workbook
        
        wb = load_workbook(report_path)
        ws = wb.create_sheet("AI Validation Log")
        
        # Заголовки
        headers = ['Маркетплейс', 'Столбец', 'Исходное значение', 'Сопоставлено с', 'Метод']
        ws.append(headers)
        
        # Данные
        for log_entry in self.ai_validation_log:
            ws.append([
                log_entry['Маркетплейс'],
                log_entry['Столбец'],
                log_entry['Исходное значение'],
                log_entry['Сопоставлено с'],
                log_entry['Метод']
            ])
        
        wb.save(report_path)
        logger.info(f"✅ Добавлен лист 'AI Validation Log' в {report_path}")
