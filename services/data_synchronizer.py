"""
–ú–æ–¥—É–ª—å –¥–ª—è —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏–∏ –¥–∞–Ω–Ω—ã—Ö –º–µ–∂–¥—É –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞–º–∏
"""
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional
from utils.excel_reader import ExcelReader
from utils.excel_writer import ExcelWriter
from config.config import FILE_CONFIGS, is_excluded_column
from services.ai_comparator import AIComparator
from utils.logger_config import setup_logger
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

# –°–æ–∑–¥–∞–µ–º –≥–ª–æ–±–∞–ª—å–Ω—ã–π –ª–æ–≥–≥–µ—Ä
logger = setup_logger('data_sync')  

class DimensionsSynchronizer:
    """–°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏—è –∫–æ–º–ø–æ–∑–∏—Ç–Ω—ã—Ö –≥–∞–±–∞—Ä–∏—Ç–æ–≤ (–î–ª–∏–Ω–∞/–®–∏—Ä–∏–Ω–∞/–í—ã—Å–æ—Ç–∞)"""
    
    # –ú–∞–ø–ø–∏–Ω–≥ —Å—Ç–æ–ª–±—Ü–æ–≤ –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞
    DIMENSIONS_MAPPING = {
        'wildberries': {
            'length': '–î–ª–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)',
            'width': '–®–∏—Ä–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)',
            'height': '–í—ã—Å–æ—Ç–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)',
            'unit': 'cm'
        },
        'ozon': {
            'length': '–î–ª–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏, –º–º*',
            'width': '–®–∏—Ä–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏, –º–º*',
            'height': '–í—ã—Å–æ—Ç–∞ —É–ø–∞–∫–æ–≤–∫–∏, –º–º*',
            'unit': 'mm'
        },
        'yandex': {
            'composite': '–ì–∞–±–∞—Ä–∏—Ç—ã —Å —É–ø–∞–∫–æ–≤–∫–æ–π, —Å–º',
            'unit': 'cm'
        }
    }
    
    @staticmethod
    def parse_composite_dimensions(value: str) -> Optional[Dict[str, float]]:
        """
        –ü–∞—Ä—Å–∏—Ç —Å—Ç—Ä–æ–∫—É "71/68/197" –≤ —Å–ª–æ–≤–∞—Ä—å {length, width, height}
        
        Args:
            value: —Å—Ç—Ä–æ–∫–∞ —Ñ–æ—Ä–º–∞—Ç–∞ "–î–ª–∏–Ω–∞/–®–∏—Ä–∏–Ω–∞/–í—ã—Å–æ—Ç–∞"
            
        Returns:
            {'length': 71.0, 'width': 68.0, 'height': 197.0} –∏–ª–∏ None
        """
        if pd.isna(value) or not str(value).strip():
            return None
        
        try:
            parts = str(value).strip().split('/')
            if len(parts) != 3:
                return None
            
            # –£–¥–∞–ª—è–µ–º –ø—Ä–æ–±–µ–ª—ã –∏ –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ float
            dimensions = {
                'length': float(parts[0].strip()),
                'width': float(parts[1].strip()),
                'height': float(parts[2].strip())
            }
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ –≤—Å–µ –∑–Ω–∞—á–µ–Ω–∏—è –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–µ
            if all(v > 0 for v in dimensions.values()):
                return dimensions
            
        except (ValueError, AttributeError):
            pass
        
        return None
    
    @staticmethod
    def format_composite_dimensions(length: float, width: float, height: float) -> str:
        """
        –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç –≥–∞–±–∞—Ä–∏—Ç—ã –≤ —Å—Ç—Ä–æ–∫—É "–î–ª–∏–Ω–∞/–®–∏—Ä–∏–Ω–∞/–í—ã—Å–æ—Ç–∞"
        
        Args:
            length: –¥–ª–∏–Ω–∞ –≤ —Å–º
            width: —à–∏—Ä–∏–Ω–∞ –≤ —Å–º  
            height: –≤—ã—Å–æ—Ç–∞ –≤ —Å–º
            
        Returns:
            –°—Ç—Ä–æ–∫–∞ —Ñ–æ—Ä–º–∞—Ç–∞ "71/68/197"
        """
        # –û–∫—Ä—É–≥–ª—è–µ–º –¥–æ —Ü–µ–ª—ã—Ö –µ—Å–ª–∏ –±–ª–∏–∑–∫–æ –∫ —Ü–µ–ª–æ–º—É, –∏–Ω–∞—á–µ –¥–æ 1 –∑–Ω–∞–∫–∞
        def smart_format(val):
            if abs(val - round(val)) < 0.01:
                return str(int(round(val)))
            return f"{val:.1f}"
        
        return f"{smart_format(length)}/{smart_format(width)}/{smart_format(height)}"
    
    @staticmethod
    def mm_to_cm(value: float) -> float:
        """–ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç –º–∏–ª–ª–∏–º–µ—Ç—Ä—ã –≤ —Å–∞–Ω—Ç–∏–º–µ—Ç—Ä—ã"""
        return value / 10
    
    @staticmethod
    def cm_to_mm(value: float) -> float:
        """–ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç —Å–∞–Ω—Ç–∏–º–µ—Ç—Ä—ã –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä—ã"""
        return value * 10
    
    @classmethod
    def sync_dimensions(cls, dfs: Dict[str, pd.DataFrame]) -> int:
        """
        –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ—Ç –≥–∞–±–∞—Ä–∏—Ç—ã –º–µ–∂–¥—É –≤—Å–µ–º–∏ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞–º–∏
        
        Args:
            dfs: —Å–ª–æ–≤–∞—Ä—å DataFrame'–æ–≤ –ø–æ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞–º
            
        Returns:
            –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
        """
        synced_count = 0
        
        # –ü–æ–ª—É—á–∞–µ–º –Ω–∞–∑–≤–∞–Ω–∏—è —Å—Ç–æ–ª–±—Ü–æ–≤ –∞—Ä—Ç–∏–∫—É–ª–æ–≤
        article_cols = {
            'wildberries': '–ê—Ä—Ç–∏–∫—É–ª –ø—Ä–æ–¥–∞–≤—Ü–∞',
            'ozon': '–ê—Ä—Ç–∏–∫—É–ª*',
            'yandex': '–í–∞—à SKU *'
        }
        
        # –°–æ–∑–¥–∞—ë–º –º–∞–ø–ø–∏–Ω–≥ –∞—Ä—Ç–∏–∫—É–ª ‚Üí –¥–∞–Ω–Ω—ã–µ
        yandex_dimensions = {}  # {article: {'length': 71, 'width': 68, 'height': 197}}
        wb_dimensions = {}
        ozon_dimensions = {}
        
        # 1. –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ –Ø–Ω–¥–µ–∫—Å (–∫–æ–º–ø–æ–∑–∏—Ç–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç)
        if 'yandex' in dfs and cls.DIMENSIONS_MAPPING['yandex']['composite'] in dfs['yandex'].columns:
            for idx, row in dfs['yandex'].iterrows():
                article = row.get(article_cols['yandex'])
                if pd.notna(article) and str(article).strip():
                    composite = row.get(cls.DIMENSIONS_MAPPING['yandex']['composite'])
                    dimensions = cls.parse_composite_dimensions(composite)
                    if dimensions:
                        yandex_dimensions[str(article).strip()] = dimensions
        
        # 2. –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ WB (—Ä–∞–∑–¥–µ–ª—å–Ω—ã–µ —Å—Ç–æ–ª–±—Ü—ã, —Å–º)
        if 'wildberries' in dfs:
            wb_map = cls.DIMENSIONS_MAPPING['wildberries']
            df_wb = dfs['wildberries']
            
            for col in [wb_map['length'], wb_map['width'], wb_map['height']]:
                if col not in df_wb.columns:
                    logger.warning(f"[WB] –°—Ç–æ–ª–±–µ—Ü '{col}' –Ω–µ –Ω–∞–π–¥–µ–Ω!")
                    return synced_count
            
            for idx, row in df_wb.iterrows():
                article = row.get(article_cols['wildberries'])
                if pd.notna(article) and str(article).strip():
                    article_str = str(article).strip()
                    length = row.get(wb_map['length'])
                    width = row.get(wb_map['width'])
                    height = row.get(wb_map['height'])
                    
                    # ‚úÖ –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ –í–°–ï –¢–†–ò –∑–Ω–∞—á–µ–Ω–∏—è –∑–∞–ø–æ–ª–Ω–µ–Ω—ã
                    if all(pd.notna(v) and str(v).strip() for v in [length, width, height]):
                        try:
                            wb_dimensions[article_str] = {
                                'length': float(length),
                                'width': float(width),
                                'height': float(height)
                            }
                        except ValueError:
                            pass
        
        # 3. –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Ozon (—Ä–∞–∑–¥–µ–ª—å–Ω—ã–µ —Å—Ç–æ–ª–±—Ü—ã, –º–º)
        if 'ozon' in dfs:
            ozon_map = cls.DIMENSIONS_MAPPING['ozon']
            df_ozon = dfs['ozon']
            
            for col in [ozon_map['length'], ozon_map['width'], ozon_map['height']]:
                if col not in df_ozon.columns:
                    logger.warning(f"[OZON] –°—Ç–æ–ª–±–µ—Ü '{col}' –Ω–µ –Ω–∞–π–¥–µ–Ω!")
                    return synced_count
            
            for idx, row in df_ozon.iterrows():
                article = row.get(article_cols['ozon'])
                if pd.notna(article) and str(article).strip():
                    article_str = str(article).strip()
                    
                    length_mm = row.get(ozon_map['length'])
                    width_mm = row.get(ozon_map['width'])
                    height_mm = row.get(ozon_map['height'])
                    
                    if all(pd.notna(v) and str(v).strip() for v in [length_mm, width_mm, height_mm]):
                        try:
                            # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –º–º ‚Üí —Å–º
                            ozon_dimensions[article_str] = {
                                'length': cls.mm_to_cm(float(length_mm)),
                                'width': cls.mm_to_cm(float(width_mm)),
                                'height': cls.mm_to_cm(float(height_mm))
                            }
                        except ValueError:
                            pass
        
        # 4. –°–ò–ù–•–†–û–ù–ò–ó–ê–¶–ò–Ø: –Ø–Ω–¥–µ–∫—Å ‚Üí WB/Ozon
        for article, dimensions in yandex_dimensions.items():
            # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏—è –≤ WB
            if 'wildberries' in dfs:
                df_wb = dfs['wildberries']
                wb_map = cls.DIMENSIONS_MAPPING['wildberries']
                
                # –ò—â–µ–º —Å—Ç—Ä–æ–∫—É —Å —ç—Ç–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º
                mask = df_wb[article_cols['wildberries']].astype(str).str.strip() == article
                if mask.any():
                    idx = df_wb[mask].index[0]
                    
                    # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –≤ —Å–º (–∫–∞–∫ –µ—Å—Ç—å)
                    if pd.isna(df_wb.at[idx, wb_map['length']]) or not str(df_wb.at[idx, wb_map['length']]).strip():
                        df_wb.at[idx, wb_map['length']] = dimensions['length']
                        synced_count += 1
                    
                    if pd.isna(df_wb.at[idx, wb_map['width']]) or not str(df_wb.at[idx, wb_map['width']]).strip():
                        df_wb.at[idx, wb_map['width']] = dimensions['width']
                        synced_count += 1
                    
                    if pd.isna(df_wb.at[idx, wb_map['height']]) or not str(df_wb.at[idx, wb_map['height']]).strip():
                        df_wb.at[idx, wb_map['height']] = dimensions['height']
                        synced_count += 1
            
            # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏—è –≤ Ozon
            if 'ozon' in dfs:
                df_ozon = dfs['ozon']
                ozon_map = cls.DIMENSIONS_MAPPING['ozon']
                
                mask = df_ozon[article_cols['ozon']].astype(str).str.strip() == article
                if mask.any():
                    idx = df_ozon[mask].index[0]
                    
                    # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –≤ –º–º (–∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –∏–∑ —Å–º)
                    if pd.isna(df_ozon.at[idx, ozon_map['length']]) or not str(df_ozon.at[idx, ozon_map['length']]).strip():
                        df_ozon.at[idx, ozon_map['length']] = int(cls.cm_to_mm(dimensions['length']))
                        synced_count += 1
                    
                    if pd.isna(df_ozon.at[idx, ozon_map['width']]) or not str(df_ozon.at[idx, ozon_map['width']]).strip():
                        df_ozon.at[idx, ozon_map['width']] = int(cls.cm_to_mm(dimensions['width']))
                        synced_count += 1
                    
                    if pd.isna(df_ozon.at[idx, ozon_map['height']]) or not str(df_ozon.at[idx, ozon_map['height']]).strip():
                        df_ozon.at[idx, ozon_map['height']] = int(cls.cm_to_mm(dimensions['height']))
                        synced_count += 1
        
        # 5. –°–ò–ù–•–†–û–ù–ò–ó–ê–¶–ò–Ø: WB ‚Üí –Ø–Ω–¥–µ–∫—Å
        for article, dimensions in wb_dimensions.items():
            if article in yandex_dimensions:
                continue  # –£–∂–µ –µ—Å—Ç—å –¥–∞–Ω–Ω—ã–µ –∏–∑ –Ø–Ω–¥–µ–∫—Å–∞
            
            if 'yandex' in dfs:
                df_yandex = dfs['yandex']
                yandex_col = cls.DIMENSIONS_MAPPING['yandex']['composite']
                
                mask = df_yandex[article_cols['yandex']].astype(str).str.strip() == article
                if mask.any():
                    idx = df_yandex[mask].index[0]
                    
                    if pd.isna(df_yandex.at[idx, yandex_col]) or not str(df_yandex.at[idx, yandex_col]).strip():
                        composite = cls.format_composite_dimensions(
                            dimensions['length'],
                            dimensions['width'],
                            dimensions['height']
                        )
                        df_yandex.at[idx, yandex_col] = composite
                        synced_count += 1
                        logger.info(f"[WB‚Üí–Ø–Ω–¥–µ–∫—Å] {article}: {composite}")
        
        # 6. –°–ò–ù–•–†–û–ù–ò–ó–ê–¶–ò–Ø: Ozon ‚Üí –Ø–Ω–¥–µ–∫—Å (–∏ –≤ WB –µ—Å–ª–∏ –Ω–µ—Ç)
        for article, dimensions in ozon_dimensions.items():
           
            
            # –í –Ø–Ω–¥–µ–∫—Å
            if 'yandex' in dfs:
                df_yandex = dfs['yandex']
                yandex_col = cls.DIMENSIONS_MAPPING['yandex']['composite']
                mask = df_yandex[article_cols['yandex']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_yandex[mask].index[0]
                    if pd.isna(df_yandex.at[idx, yandex_col]) or not str(df_yandex.at[idx, yandex_col]).strip():
                        composite = cls.format_composite_dimensions(
                            dimensions['length'],
                            dimensions['width'],
                            dimensions['height']
                        )
                        df_yandex.at[idx, yandex_col] = composite
                        synced_count += 1
                        logger.info(f"[Ozon‚Üí–Ø–Ω–¥–µ–∫—Å] {article}: {composite}")
            
            # –í WB (—Ç–µ–ø–µ—Ä—å —ç—Ç–æ –±—É–¥–µ—Ç —Ä–∞–±–æ—Ç–∞—Ç—å –≤—Å–µ–≥–¥–∞!)
            if 'wildberries' in dfs:
                df_wb = dfs['wildberries']
                wb_map = cls.DIMENSIONS_MAPPING['wildberries']
                mask = df_wb[article_cols['wildberries']].astype(str).str.strip() == article
                
                if mask.any():
                    idx = df_wb[mask].index[0]
                    
                    # –ó–∞–ø–æ–ª–Ω—è–µ–º –¢–û–õ–¨–ö–û –ø—É—Å—Ç—ã–µ –ø–æ–ª—è
                    if pd.isna(df_wb.at[idx, wb_map['length']]) or not str(df_wb.at[idx, wb_map['length']]).strip():
                        df_wb.at[idx, wb_map['length']] = dimensions['length']
                        synced_count += 1
                        logger.info(f"[Ozon‚ÜíWB] {article}: length={dimensions['length']}")
                    
                    if pd.isna(df_wb.at[idx, wb_map['width']]) or not str(df_wb.at[idx, wb_map['width']]).strip():
                        df_wb.at[idx, wb_map['width']] = dimensions['width']
                        synced_count += 1
                        logger.info(f"[Ozon‚ÜíWB] {article}: width={dimensions['width']}")
                    
                    if pd.isna(df_wb.at[idx, wb_map['height']]) or not str(df_wb.at[idx, wb_map['height']]).strip():
                        df_wb.at[idx, wb_map['height']] = dimensions['height']
                        synced_count += 1
                        logger.info(f"[Ozon‚ÜíWB] {article}: height={dimensions['height']}")
        
        logger.info(f"‚úÖ –ì–∞–±–∞—Ä–∏—Ç—ã: —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä–æ–≤–∞–Ω–æ {synced_count} –∑–Ω–∞—á–µ–Ω–∏–π")
        return synced_count


class DataSynchronizer:
    """–ö–ª–∞—Å—Å –¥–ª—è —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏–∏ –¥–∞–Ω–Ω—ã—Ö –º–µ–∂–¥—É —Ç—Ä–µ–º—è –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞–º–∏"""
    
    def __init__(self, comparison_result: Dict, ai_comparator=None):
        self.comparison_result = comparison_result
        self.article_columns = {
            'wildberries': '–ê—Ä—Ç–∏–∫—É–ª –ø—Ä–æ–¥–∞–≤—Ü–∞',
            'ozon': '–ê—Ä—Ç–∏–∫—É–ª*',
            'yandex': '–í–∞—à SKU *'
        }
        self.changes_log = {
            'wildberries': [],
            'ozon': [],
            'yandex': []
        }
        self.original_file_paths = {}
        self.ai_comparator = ai_comparator
        
        
        # –î–û–ë–ê–í–¨–¢–ï: –ö—ç—à validation –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Å—Ç–æ–ª–±—Ü–∞
        self.column_validations = {}  # {marketplace: {column_name: [allowed_values]}}
        self.original_column_names = {}
        # –î–û–ë–ê–í–¨–¢–ï –≠–¢–£ –°–¢–†–û–ö–£:
        self.ai_validation_log = []  # –õ–æ–≥–∏ AI-—Å–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–∏–π
        logger.info("–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è DataSynchronizer")
        logger.debug(f"AI comparator –ø–µ—Ä–µ–¥–∞–Ω: {ai_comparator is not None}")
    
    def _align_articles(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """
        –í—ã—Ä–∞–≤–Ω–∏–≤–∞–µ—Ç –∞—Ä—Ç–∏–∫—É–ª—ã –º–µ–∂–¥—É –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞–º–∏ - –¥–æ–±–∞–≤–ª—è–µ—Ç –æ—Ç—Å—É—Ç—Å—Ç–≤—É—é—â–∏–µ —Å—Ç—Ä–æ–∫–∏
        Args:
            dfs: —Å–ª–æ–≤–∞—Ä—å —Å DataFrame –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞
        Returns:
            –û–±–Ω–æ–≤–ª–µ–Ω–Ω—ã–µ DataFrame —Å –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–º–∏ –∞—Ä—Ç–∏–∫—É–ª–∞–º–∏
        """
        logger.info("\n" + "="*60)
        logger.info("–í–´–†–ê–í–ù–ò–í–ê–ù–ò–ï –ê–†–¢–ò–ö–£–õ–û–í –ú–ï–ñ–î–£ –ú–ê–†–ö–ï–¢–ü–õ–ï–ô–°–ê–ú–ò")
        logger.info("="*60)
        
        # –°–æ–±–∏—Ä–∞–µ–º –≤—Å–µ —É–Ω–∏–∫–∞–ª—å–Ω—ã–µ –∞—Ä—Ç–∏–∫—É–ª—ã –∏–∑ –≤—Å–µ—Ö –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤
        all_articles = set()
        for marketplace in ['wildberries', 'ozon', 'yandex']:
            article_col = self.article_columns[marketplace]
            if article_col in dfs[marketplace].columns:
                articles = dfs[marketplace][article_col].dropna().astype(str).str.strip()
                articles = articles[articles != '']  # –£–±–∏—Ä–∞–µ–º –ø—É—Å—Ç—ã–µ
                
                # –§–∏–ª—å—Ç—Ä–∞—Ü–∏—è: –£–±–∏—Ä–∞–µ–º –æ–ø–∏—Å–∞–Ω–∏—è –ø–æ–ª–µ–π –∏ —Å–ª–∏—à–∫–æ–º –¥–ª–∏–Ω–Ω—ã–µ —Å—Ç—Ä–æ–∫–∏
                articles = articles[
                    ~articles.str.contains(
                        '–∏–¥–µ–Ω—Ç–∏—Ñ–∏—Ü–∏—Ä–æ–≤–∞—Ç—å|–æ–ø–∏—Å–∞–Ω–∏–µ|–∑–∞–ø–æ–ª–Ω–∏—Ç—å|–ø—Ä–∏–º–µ—Ä|–Ω–∞–∑–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞|–ø–æ –∫–æ—Ç–æ—Ä–æ–º—É',
                        case=False,
                        na=False
                    )
                ]
                # –£–±–∏—Ä–∞–µ–º —Å—Ç—Ä–æ–∫–∏ –¥–ª–∏–Ω–Ω–µ–µ 50 —Å–∏–º–≤–æ–ª–æ–≤ (—Å–∫–æ—Ä–µ–µ –≤—Å–µ–≥–æ –æ–ø–∏—Å–∞–Ω–∏–µ)
                articles = articles[articles.str.len() < 50]
                all_articles.update(articles.tolist())
                logger.info(f"üìä {marketplace.upper()}: {len(articles)} –∞—Ä—Ç–∏–∫—É–ª–æ–≤")
        
        logger.info(f"\nüîç –í—Å–µ–≥–æ —É–Ω–∏–∫–∞–ª—å–Ω—ã—Ö –∞—Ä—Ç–∏–∫—É–ª–æ–≤: {len(all_articles)}")
        
        # –î–ª—è –∫–∞–∂–¥–æ–≥–æ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞ –ø—Ä–æ–≤–µ—Ä—è–µ–º –Ω–µ–¥–æ—Å—Ç–∞—é—â–∏–µ –∞—Ä—Ç–∏–∫—É–ª—ã
        total_added = 0
        for marketplace in ['wildberries', 'ozon', 'yandex']:
            article_col = self.article_columns[marketplace]
            if article_col not in dfs[marketplace].columns:
                logger.warning(f"‚ö†Ô∏è {marketplace.upper()}: —Å—Ç–æ–ª–±–µ—Ü '{article_col}' –Ω–µ –Ω–∞–π–¥–µ–Ω, –ø—Ä–æ–ø—É—Å–∫–∞—é")
                continue
            
            # –°–±—Ä–∞—Å—ã–≤–∞–µ–º –∏–Ω–¥–µ–∫—Å—ã –ü–ï–†–ï–î –æ–±—Ä–∞–±–æ—Ç–∫–æ–π!
            dfs[marketplace] = dfs[marketplace].reset_index(drop=True)
            
            # –°—É—â–µ—Å—Ç–≤—É—é—â–∏–µ –∞—Ä—Ç–∏–∫—É–ª—ã
            df = dfs[marketplace]
            
            # üÜï –ù–û–í–´–ô –ü–û–î–•–û–î: –ù–∞—Ö–æ–¥–∏–º —Å—Ç—Ä–æ–∫–∏ —Å –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã–º–∏ –∞—Ä—Ç–∏–∫—É–ª–∞–º–∏ –Ω–∞–ø—Ä—è–º—É—é –≤ DataFrame
            article_series = df[article_col].dropna().astype(str).str.strip()
            article_series = article_series[article_series != '']
            
            # –§–∏–ª—å—Ç—Ä–∞—Ü–∏—è
            valid_mask = (
                ~article_series.str.contains(
                    '–∏–¥–µ–Ω—Ç–∏—Ñ–∏—Ü–∏—Ä–æ–≤–∞—Ç—å|–æ–ø–∏—Å–∞–Ω–∏–µ|–∑–∞–ø–æ–ª–Ω–∏—Ç—å|–ø—Ä–∏–º–µ—Ä|–Ω–∞–∑–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞|–ø–æ –∫–æ—Ç–æ—Ä–æ–º—É',
                    case=False,
                    na=False
                ) & 
                (article_series.str.len() < 50)
            )
            article_series = article_series[valid_mask]
            
            # üÜï –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –ü–æ–ª—É—á–∞–µ–º –ü–û–ó–ò–¶–ò–û–ù–ù–´–ô –∏–Ω–¥–µ–∫—Å –ø–æ—Å–ª–µ–¥–Ω–µ–π –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω–æ–π —Å—Ç—Ä–æ–∫–∏
            if len(article_series) > 0:
                # –ü–æ–ª—É—á–∞–µ–º label –∏–Ω–¥–µ–∫—Å –ø–æ—Å–ª–µ–¥–Ω–µ–π –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω–æ–π —Å—Ç—Ä–æ–∫–∏
                last_label_idx = article_series.index[-1]
                # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ –ø–æ–∑–∏—Ü–∏–æ–Ω–Ω—ã–π –∏–Ω–¥–µ–∫—Å
                last_filled_position = df.index.get_loc(last_label_idx)
            else:
                last_filled_position = -1
            
            existing_articles_set = set(article_series.tolist())
            
            # –ù–∞—Ö–æ–¥–∏–º –Ω–µ–¥–æ—Å—Ç–∞—é—â–∏–µ
            missing_articles = all_articles - existing_articles_set
            
            if not missing_articles:
                logger.info(f"‚úÖ {marketplace.upper()}: –≤—Å–µ –∞—Ä—Ç–∏–∫—É–ª—ã –ø—Ä–∏—Å—É—Ç—Å—Ç–≤—É—é—Ç")
                continue
            
            logger.info(f"\n‚ûï {marketplace.upper()}: –¥–æ–±–∞–≤–ª—è—é {len(missing_articles)} –∞—Ä—Ç–∏–∫—É–ª–æ–≤")
            
            # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–µ —Å—Ç—Ä–æ–∫–∏ –¥–ª—è –Ω–µ–¥–æ—Å—Ç–∞—é—â–∏—Ö –∞—Ä—Ç–∏–∫—É–ª–æ–≤
            new_rows = []
            for article in sorted(missing_articles):
                # –°–æ–∑–¥–∞–µ–º –ø—É—Å—Ç—É—é —Å—Ç—Ä–æ–∫—É —Å–æ –≤—Å–µ–º–∏ —Å—Ç–æ–ª–±—Ü–∞–º–∏
                new_row = {col: None for col in df.columns}
                # –ó–∞–ø–æ–ª–Ω—è–µ–º —Ç–æ–ª—å–∫–æ –∞—Ä—Ç–∏–∫—É–ª
                new_row[article_col] = article
                new_rows.append(new_row)
            
            # –í—Å—Ç–∞–≤–ª—è–µ–º –Ω–æ–≤—ã–µ —Å—Ç—Ä–æ–∫–∏ –°–†–ê–ó–£ –ü–û–°–õ–ï –ø–æ—Å–ª–µ–¥–Ω–µ–π –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω–æ–π!
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                
                if last_filled_position >= 0:
                    # –ï—Å—Ç—å –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã–µ —Å—Ç—Ä–æ–∫–∏ - –≤—Å—Ç–∞–≤–ª—è–µ–º –ø–æ—Å–ª–µ –Ω–∏—Ö
                    # –ò—Å–ø–æ–ª—å–∑—É–µ–º –ü–û–ó–ò–¶–ò–û–ù–ù–´–ô –∏–Ω–¥–µ–∫—Å!
                    before = df.iloc[:last_filled_position + 1].copy()
                    after = df.iloc[last_filled_position + 1:].copy()
                    
                    # –°–∫–ª–µ–∏–≤–∞–µ–º: –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã–µ + –Ω–æ–≤—ã–µ + –ø—É—Å—Ç—ã–µ
                    dfs[marketplace] = pd.concat([before, new_df, after], ignore_index=True)
                    
                    logger.info(f" ‚úì –î–æ–±–∞–≤–ª–µ–Ω–æ {len(new_rows)} —Å—Ç—Ä–æ–∫ –ø–æ—Å–ª–µ –ø–æ–∑–∏—Ü–∏–∏ {last_filled_position}")
                else:
                    # –ù–µ—Ç –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫ - –¥–æ–±–∞–≤–ª—è–µ–º –≤ –Ω–∞—á–∞–ª–æ
                    dfs[marketplace] = pd.concat([new_df, df], ignore_index=True)
                    logger.info(f" ‚úì –î–æ–±–∞–≤–ª–µ–Ω–æ {len(new_rows)} —Å—Ç—Ä–æ–∫ –≤ –Ω–∞—á–∞–ª–æ")
                
                total_added += len(new_rows)
                logger.info(f" üìä –ë—ã–ª–æ: {len(df)}, —Å—Ç–∞–ª–æ: {len(dfs[marketplace])}")
        
        if total_added > 0:
            logger.info(f"\n‚úÖ –ò—Ç–æ–≥–æ –¥–æ–±–∞–≤–ª–µ–Ω–æ {total_added} –Ω–æ–≤—ã—Ö —Å—Ç—Ä–æ–∫ –≤–æ –≤—Å–µ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å—ã")
        else:
            logger.info(f"\n‚úÖ –í—ã—Ä–∞–≤–Ω–∏–≤–∞–Ω–∏–µ –Ω–µ —Ç—Ä–µ–±—É–µ—Ç—Å—è - –≤—Å–µ –∞—Ä—Ç–∏–∫—É–ª—ã –ø—Ä–∏—Å—É—Ç—Å—Ç–≤—É—é—Ç")
        
        return dfs
    
    
    def _detect_unit(self, column_name: str) -> Optional[str]:
        """
        –û–ø—Ä–µ–¥–µ–ª—è–µ—Ç –µ–¥–∏–Ω–∏—Ü—É –∏–∑–º–µ—Ä–µ–Ω–∏—è –∏–∑ –Ω–∞–∑–≤–∞–Ω–∏—è —Å—Ç–æ–ª–±—Ü–∞
        
        Args:
            column_name: –Ω–∞–∑–≤–∞–Ω–∏–µ —Å—Ç–æ–ª–±—Ü–∞
        
        Returns:
            –ï–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è ('kg', 'g', 'mm', 'cm') –∏–ª–∏ None
        """
        if not column_name:
            return None
        
        column_lower = column_name.lower()
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –µ–¥–∏–Ω–∏—Ü—ã –≤–µ—Å–∞
        if '–∫–≥' in column_lower or 'kg' in column_lower:
            return 'kg'
        if ' –≥' in column_lower or ',–≥' in column_lower or 'gram' in column_lower or column_lower.endswith('–≥'):
            return 'g'
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –µ–¥–∏–Ω–∏—Ü—ã –¥–ª–∏–Ω—ã/—Ä–∞–∑–º–µ—Ä–∞
        if '–º–º' in column_lower or 'mm' in column_lower:
            return 'mm'
        if '—Å–º' in column_lower or 'cm' in column_lower:
            return 'cm'
        
        return None
    
    def _convert_value(
        self, 
        value, 
        from_unit: Optional[str], 
        to_unit: Optional[str]
    ):
        """
        –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç –∑–Ω–∞—á–µ–Ω–∏–µ –º–µ–∂–¥—É –µ–¥–∏–Ω–∏—Ü–∞–º–∏ –∏–∑–º–µ—Ä–µ–Ω–∏—è
        
        Args:
            value: –∏—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
            from_unit: –∏—Å—Ö–æ–¥–Ω–∞—è –µ–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è
            to_unit: —Ü–µ–ª–µ–≤–∞—è –µ–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è
        
        Returns:
            –°–∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
        """
        # –ï—Å–ª–∏ –µ–¥–∏–Ω–∏—Ü—ã –∏–∑–º–µ—Ä–µ–Ω–∏—è –Ω–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω—ã –∏–ª–∏ –æ–¥–∏–Ω–∞–∫–æ–≤—ã–µ - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∫–∞–∫ –µ—Å—Ç—å
        if not from_unit or not to_unit or from_unit == to_unit:
            return value
        
        # –ï—Å–ª–∏ –∑–Ω–∞—á–µ–Ω–∏–µ –ø—É—Å—Ç–æ–µ –∏–ª–∏ –Ω–µ —á–∏—Å–ª–æ–≤–æ–µ - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∫–∞–∫ –µ—Å—Ç—å
        if pd.isna(value):
            return value
        
        try:
            numeric_value = float(value)
        except (ValueError, TypeError):
            return value
        
        # –ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è –≤–µ—Å–∞
        if from_unit == 'kg' and to_unit == 'g':
            result = numeric_value * 1000
            print(f"      [–ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è] {numeric_value} –∫–≥ ‚Üí {result} –≥")
            return result
        elif from_unit == 'g' and to_unit == 'kg':
            result = numeric_value / 1000
            print(f"      [–ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è] {numeric_value} –≥ ‚Üí {result} –∫–≥")
            return result
        
        # –ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è —Ä–∞–∑–º–µ—Ä–æ–≤
        elif from_unit == 'mm' and to_unit == 'cm':
            result = numeric_value / 10
            print(f"      [–ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è] {numeric_value} –º–º ‚Üí {result} —Å–º")
            return result
        elif from_unit == 'cm' and to_unit == 'mm':
            result = numeric_value * 10
            print(f"      [–ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è] {numeric_value} —Å–º ‚Üí {result} –º–º")
            return result
        
        # –ï—Å–ª–∏ –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç—Å—è - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∏—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
        return value
    
    def synchronize_data(
        self,
        file_paths: Dict[str, str],
        output_paths: Dict[str, str] = None,
        report_path: str = None  # ‚Üê –î–û–ë–ê–í–ò–¢–¨
    ) -> Tuple[Dict[str, pd.DataFrame], Dict]:
        logger.info("="*60)
        logger.info("–°–ò–ù–•–†–û–ù–ò–ó–ê–¶–ò–Ø –î–ê–ù–ù–´–• –ú–ï–ñ–î–£ –ú–ê–†–ö–ï–¢–ü–õ–ï–ô–°–ê–ú–ò")
        logger.info("="*60)
        
        # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ –≤—Å–µ—Ö —Ç—Ä–µ—Ö —Ñ–∞–π–ª–æ–≤
        dfs = self._load_all_dataframes(file_paths)

        # üÜï 3. –°–ò–ù–•–†–û–ù–ò–ó–ê–¶–ò–Ø –ö–û–ú–ü–û–ó–ò–¢–ù–´–• –ì–ê–ë–ê–†–ò–¢–û–í (–ù–û–í–û–ï!)
        logger.info("\n[*] –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏—è –≥–∞–±–∞—Ä–∏—Ç–æ–≤...")
        dimensions_synced = DimensionsSynchronizer.sync_dimensions(dfs)
        
        # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ
        synced_dfs = self._sync_all_matches(dfs)

        # ‚≠ê –ü–û–°–¢–û–ë–†–ê–ë–û–¢–ö–ê: –ö–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è –≥–∞–±–∞—Ä–∏—Ç–æ–≤ WB (–º–º ‚Üí —Å–º –µ—Å–ª–∏ –∏–∑ Ozon)
        logger.info("\n[*] –ü–æ—Å—Ç–æ–±—Ä–∞–±–æ—Ç–∫–∞ –≥–∞–±–∞—Ä–∏—Ç–æ–≤ WB...")
        converted_count = self._postprocess_wb_dimensions(synced_dfs)
        if converted_count > 0:
            logger.info(f"‚úÖ –°–∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–æ {converted_count} –∑–Ω–∞—á–µ–Ω–∏–π –≥–∞–±–∞—Ä–∏—Ç–æ–≤ (–º–º ‚Üí —Å–º)")

        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
        if output_paths:
            self._save_results(synced_dfs, output_paths)
        
        logger.info("‚úÖ –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏—è –∑–∞–≤–µ—Ä—à–µ–Ω–∞!")
        
        
        return synced_dfs, self.changes_log
    
    def _load_all_dataframes(self, file_paths: Dict[str, str]) -> Dict[str, pd.DataFrame]:
        """–ó–∞–≥—Ä—É–∂–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ —á–µ—Ä–µ–∑ openpyxl –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ñ–æ—Ä–º–∞—Ç–æ–≤"""
        logger.info("üìÇ –ó–∞–≥—Ä—É–∂–∞—é –¥–∞–Ω–Ω—ã–µ –∏–∑ —Ñ–∞–π–ª–æ–≤...")
        
        dfs = {}
        # üÜï –°–ª–æ–≤–∞—Ä—å –¥–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω—ã—Ö –Ω–∞–∑–≤–∞–Ω–∏–π —Å—Ç–æ–ª–±—Ü–æ–≤
        self.original_column_names = {}
        
        for marketplace, file_path in file_paths.items():
            self.original_file_paths[marketplace] = file_path
            config = FILE_CONFIGS[marketplace]
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb[config['sheet_name']]
            
            self._load_column_validations(ws, marketplace, config)
            
            data = []
            headers = []
            
            # –ß–∏—Ç–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏
            for cell in ws[config['header_row']]:
                headers.append(cell.value if cell.value else '')
            
            # üÜï –û–ë–†–ê–ë–û–¢–ö–ê –î–£–ë–õ–ò–ö–ê–¢–û–í –°–¢–û–õ–ë–¶–û–í
            original_headers = headers.copy()
            seen = {}
            renamed_columns = {}
            
            for i, col in enumerate(headers):
                if col in seen:
                    # –ù–∞—à–ª–∏ –¥—É–±–ª–∏–∫–∞—Ç - –¥–æ–±–∞–≤–ª—è–µ–º —Å—É—Ñ—Ñ–∏–∫—Å
                    seen[col] += 1
                    new_name = f"{col}{seen[col]}"
                    logger.warning(f"‚ö†Ô∏è [{marketplace}] –î—É–±–ª–∏–∫–∞—Ç —Å—Ç–æ–ª–±—Ü–∞ '{col}' –ø–µ—Ä–µ–∏–º–µ–Ω–æ–≤–∞–Ω –≤ '{new_name}'")
                    headers[i] = new_name
                    renamed_columns[new_name] = col  # –°–æ—Ö—Ä–∞–Ω—è–µ–º –º–∞–ø–ø–∏–Ω–≥ –¥–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞
                else:
                    seen[col] = 0
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –¥–ª—è –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏—è
            if renamed_columns:
                self.original_column_names[marketplace] = {
                    'renamed': renamed_columns,
                    'all_headers': original_headers
                }
            
            # –ò—Å–ø–æ–ª—å–∑—É–µ–º data_start_row –≤–º–µ—Å—Ç–æ header_row + 1
            data_start = config.get('data_start_row', config['header_row'] + 1)
            
            # –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ
            for row in ws.iter_rows(min_row=data_start, values_only=True):
                data.append(row)
            
            df = pd.DataFrame(data, columns=headers)
            dfs[marketplace] = df
            
            wb.close()
            logger.info(f"‚úÖ {config['display_name']}: –∑–∞–≥—Ä—É–∂–µ–Ω–æ {len(df)} —Ç–æ–≤–∞—Ä–æ–≤")
        
        
        return dfs


    def _load_column_validations(self, ws, marketplace: str, config: Dict):
        """
        –ó–∞–≥—Ä—É–∂–∞–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ validation –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Å—Ç–æ–ª–±—Ü–∞
        """
        from openpyxl.utils import range_boundaries
        
        if marketplace not in self.column_validations:
            self.column_validations[marketplace] = {}
        
        header_row = config['header_row']
        
        # –°–æ–∑–¥–∞–µ–º –º–∞–ø–ø–∏–Ω–≥: –Ω–æ–º–µ—Ä –∫–æ–ª–æ–Ω–∫–∏ -> –Ω–∞–∑–≤–∞–Ω–∏–µ
        col_idx_to_name = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                col_name = str(cell.value).strip()
                col_idx_to_name[col_idx] = col_name
        
        logger.info(f"üìã [{marketplace}] –ù–∞–π–¥–µ–Ω–æ {len(col_idx_to_name)} —Å—Ç–æ–ª–±—Ü–æ–≤")
        logger.debug(f"[{marketplace}] –ü–µ—Ä–≤—ã–µ 5 —Å—Ç–æ–ª–±—Ü–æ–≤: {list(col_idx_to_name.values())[:5]}")
        
        # –î–û–ë–ê–í–¨–¢–ï: –ü–æ–ª—É—á–∞–µ–º –∏–º–µ–Ω–æ–≤–∞–Ω–Ω—ã–µ –¥–∏–∞–ø–∞–∑–æ–Ω—ã
        workbook = ws.parent
        named_ranges = {}
        
        try:
            # –ü–†–ê–í–ò–õ–¨–ù–´–ô —Å–ø–æ—Å–æ–± –∏—Ç–µ—Ä–∞—Ü–∏–∏ –ø–æ –∏–º–µ–Ω–æ–≤–∞–Ω–Ω—ã–º –¥–∏–∞–ø–∞–∑–æ–Ω–∞–º
            for name_obj in workbook.defined_names.values():
                try:
                    if name_obj.value:
                        named_ranges[name_obj.name] = name_obj.value
                except Exception as e:
                    logger.debug(f"[{marketplace}] –ü—Ä–æ–ø—É—â–µ–Ω –∏–º–µ–Ω–æ–≤–∞–Ω–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω: {e}")
            
            logger.info(f"[{marketplace}] –ù–∞–π–¥–µ–Ω–æ {len(named_ranges)} –∏–º–µ–Ω–æ–≤–∞–Ω–Ω—ã—Ö –¥–∏–∞–ø–∞–∑–æ–Ω–æ–≤")
            
            # –í—ã–≤–æ–¥–∏–º –ø–µ—Ä–≤—ã–µ 5 –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏
            if named_ranges:
                sample = list(named_ranges.items())[:5]
                for name, value in sample:
                    logger.debug(f"[{marketplace}] –ò–º–µ–Ω–æ–≤–∞–Ω–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω '{name}' = '{value}'")
        except Exception as e:
            logger.error(f"[{marketplace}] –û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –∏–º–µ–Ω–æ–≤–∞–Ω–Ω—ã—Ö –¥–∏–∞–ø–∞–∑–æ–Ω–æ–≤: {e}")
        
        # –ü—Ä–æ—Ö–æ–¥–∏–º –ø–æ –≤—Å–µ–º validation –ø—Ä–∞–≤–∏–ª–∞–º
        validation_count = 0
        dv_index = 0
        
        for dv in ws.data_validations.dataValidation:
            dv_index += 1
            logger.debug(f"[{marketplace}] DV #{dv_index}: type={dv.type}, sqref={dv.sqref}")
            
            if dv.type != "list":
                logger.debug(f"[{marketplace}] DV #{dv_index}: –ü–†–û–ü–£–©–ï–ù (type != 'list')")
                continue
                
            if dv.sqref is None:
                logger.debug(f"[{marketplace}] DV #{dv_index}: –ü–†–û–ü–£–©–ï–ù (sqref is None)")
                continue
            
            # –ò–∑–≤–ª–µ–∫–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è –∏–∑ validation
            allowed_values = []
            if dv.formula1:
                formula = dv.formula1
                logger.debug(f"[{marketplace}] DV #{dv_index}: formula1='{formula[:100]}'...")
                
                # –°–ø–∏—Å–æ–∫ –∑–∞–¥–∞–Ω –ø—Ä—è–º–æ: "–ö—Ä–∞—Å–Ω—ã–π,–°–∏–Ω–∏–π,–ó–µ–ª–µ–Ω—ã–π"
                if formula.startswith('"') and formula.endswith('"'):
                    allowed_values = [v.strip() for v in formula.strip('"').split(',')]
                    logger.debug(f"[{marketplace}] DV #{dv_index}: –ü—Ä—è–º–æ–π —Å–ø–∏—Å–æ–∫, {len(allowed_values)} –∑–Ω–∞—á–µ–Ω–∏–π")
                
                # –î–û–ë–ê–í–¨–¢–ï: –ü—Ä–æ–≤–µ—Ä—è–µ–º –∏–º–µ–Ω–æ–≤–∞–Ω–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω
                elif formula in named_ranges:
                    logger.debug(f"[{marketplace}] DV #{dv_index}: –ò–º–µ–Ω–æ–≤–∞–Ω–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω '{formula}'")
                    try:
                        range_formula = named_ranges[formula]
                        logger.debug(f"[{marketplace}] DV #{dv_index}: –§–æ—Ä–º—É–ª–∞ –¥–∏–∞–ø–∞–∑–æ–Ω–∞: '{range_formula}'")
                        
                        # –ü–∞—Ä—Å–∏–º —Ñ–æ—Ä–º—É–ª—É –≤–∏–¥–∞ "–õ–∏—Å—Ç!$A$1:$A$10"
                        clean_formula = range_formula.replace('$', '')
                        if '!' in clean_formula:
                            sheet_name, range_ref = clean_formula.split('!', 1)
                            # –£–±–∏—Ä–∞–µ–º –∫–∞–≤—ã—á–∫–∏ –∏–∑ –∏–º–µ–Ω–∏ –ª–∏—Å—Ç–∞ –µ—Å–ª–∏ –µ—Å—Ç—å
                            sheet_name = sheet_name.strip("'")
                            target_ws = workbook[sheet_name]
                        else:
                            range_ref = clean_formula
                            target_ws = ws
                        
                        # –ò–∑–≤–ª–µ–∫–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è
                        for row in target_ws[range_ref]:
                            for cell in row:
                                if cell.value is not None:
                                    allowed_values.append(str(cell.value).strip())
                        
                        logger.info(f"‚úÖ [{marketplace}] DV #{dv_index}: –ò–∑–≤–ª–µ—á–µ–Ω–æ {len(allowed_values)} –∑–Ω–∞—á–µ–Ω–∏–π –∏–∑ –∏–º–µ–Ω–æ–≤–∞–Ω–Ω–æ–≥–æ –¥–∏–∞–ø–∞–∑–æ–Ω–∞ '{formula}'")
                    except Exception as e:
                        logger.error(f"[{marketplace}] DV #{dv_index}: –û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∏–º–µ–Ω–æ–≤–∞–Ω–Ω–æ–≥–æ –¥–∏–∞–ø–∞–∑–æ–Ω–∞ '{formula}': {e}")
                
                # –°–ø–∏—Å–æ–∫ –∑–∞–¥–∞–Ω —á–µ—Ä–µ–∑ –æ–±—ã—á–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω
                elif ':' in formula:
                    try:
                        clean_formula = formula.replace('$', '')
                        if '!' in clean_formula:
                            sheet_name, range_ref = clean_formula.split('!')
                            target_ws = workbook[sheet_name]
                            logger.debug(f"[{marketplace}] DV #{dv_index}: –î–∏–∞–ø–∞–∑–æ–Ω –Ω–∞ –ª–∏—Å—Ç–µ '{sheet_name}': {range_ref}")
                        else:
                            range_ref = clean_formula
                            target_ws = ws
                            logger.debug(f"[{marketplace}] DV #{dv_index}: –î–∏–∞–ø–∞–∑–æ–Ω –Ω–∞ —Ç–µ–∫—É—â–µ–º –ª–∏—Å—Ç–µ: {range_ref}")
                        
                        for row in target_ws[range_ref]:
                            for cell in row:
                                if cell.value is not None:
                                    allowed_values.append(str(cell.value).strip())
                        
                        logger.debug(f"[{marketplace}] DV #{dv_index}: –ò–∑–≤–ª–µ—á–µ–Ω–æ {len(allowed_values)} –∑–Ω–∞—á–µ–Ω–∏–π")
                    except Exception as e:
                        logger.error(f"[{marketplace}] DV #{dv_index}: –û—à–∏–±–∫–∞ –∏–∑–≤–ª–µ—á–µ–Ω–∏—è validation: {e}")
                else:
                    logger.warning(f"[{marketplace}] DV #{dv_index}: –ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç —Ñ–æ—Ä–º—É–ª—ã: '{formula}'")
            else:
                logger.debug(f"[{marketplace}] DV #{dv_index}: formula1 –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç")
            
            if not allowed_values:
                logger.debug(f"[{marketplace}] DV #{dv_index}: –ü–†–û–ü–£–©–ï–ù (–ø—É—Å—Ç–æ–π —Å–ø–∏—Å–æ–∫ –∑–Ω–∞—á–µ–Ω–∏–π)")
                continue
            
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–∞–∫–∏–µ —Å—Ç–æ–ª–±—Ü—ã –∑–∞—Ç—Ä–æ–Ω—É—Ç—ã —ç—Ç–∏–º validation
            ranges = str(dv.sqref).split()
            logger.debug(f"[{marketplace}] DV #{dv_index}: sqref —Å–æ–¥–µ—Ä–∂–∏—Ç {len(ranges)} –¥–∏–∞–ø–∞–∑–æ–Ω(–æ–≤): {ranges}")
            
            for range_str in ranges:
                try:
                    if ':' in range_str:
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        logger.debug(f"[{marketplace}] DV #{dv_index}: –î–∏–∞–ø–∞–∑–æ–Ω {range_str} -> —Å—Ç–æ–ª–±—Ü—ã {min_col}-{max_col}, —Å—Ç—Ä–æ–∫–∏ {min_row}-{max_row}")
                        
                        # –ü—Ä–∏–º–µ–Ω—è–µ–º validation –∫–æ –≤—Å–µ–º –∫–æ–ª–æ–Ω–∫–∞–º –≤ –¥–∏–∞–ø–∞–∑–æ–Ω–µ
                        for col_idx in range(min_col, max_col + 1):
                            if col_idx in col_idx_to_name:
                                col_name = col_idx_to_name[col_idx]
                                self.column_validations[marketplace][col_name] = allowed_values
                                validation_count += 1
                                logger.info(f"‚úÖ [{marketplace}] Validation –¥–ª—è '{col_name}': {len(allowed_values)} –∑–Ω–∞—á–µ–Ω–∏–π")
                            else:
                                logger.debug(f"[{marketplace}] DV #{dv_index}: –°—Ç–æ–ª–±–µ—Ü {col_idx} –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ –∑–∞–≥–æ–ª–æ–≤–∫–∞—Ö")
                except Exception as e:
                    logger.error(f"[{marketplace}] DV #{dv_index}: –û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ range_str '{range_str}': {e}")
        
        # –ò—Ç–æ–≥–æ–≤–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        logger.info(f"üìä [{marketplace}] –ò—Ç–æ–≥–æ –∑–∞–≥—Ä—É–∂–µ–Ω–æ validation –¥–ª—è {validation_count} —Å—Ç–æ–ª–±—Ü–æ–≤ –∏–∑ {len(ws.data_validations.dataValidation)} –ø—Ä–∞–≤–∏–ª")
        
        if validation_count == 0:
            logger.warning(f"‚ö†Ô∏è [{marketplace}] –ù–ï –ù–ê–ô–î–ï–ù–û –Ω–∏ –æ–¥–Ω–æ–≥–æ validation!")
        
        for col_name, values in self.column_validations.get(marketplace, {}).items():
            logger.debug(f"  ‚Ä¢ {col_name}: {len(values)} –∑–Ω–∞—á–µ–Ω–∏–π (–ø–µ—Ä–≤—ã–µ 3: {values[:3]})")

    
    def _sync_all_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """–°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ—Ç –≤—Å–µ —Å–æ–≤–ø–∞–¥–∞—é—â–∏–µ —Å—Ç–æ–ª–±—Ü—ã"""
        
        # –°–æ–∑–¥–∞–µ–º –∫–æ–ø–∏–∏ –¥–ª—è —Ä–∞–±–æ—Ç—ã
        synced_dfs = {
            'wildberries': dfs['wildberries'].copy(),
            'ozon': dfs['ozon'].copy(),
            'yandex': dfs['yandex'].copy()
        }
        
        # üÜï –ù–û–í–û–ï: –í—ã—Ä–∞–≤–Ω–∏–≤–∞–µ–º –∞—Ä—Ç–∏–∫—É–ª—ã –ü–ï–†–ï–î —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏–µ–π
        synced_dfs = self._align_articles(synced_dfs)
        
        # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ–º —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è –≤—Å–µ—Ö —Ç—Ä–µ—Ö –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤
        print("\n[*] –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É—é —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è –≤—Å–µ—Ö 3 –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤...")
        synced_dfs = self._sync_three_way_matches(synced_dfs)
        
        # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ–º —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è –º–µ–∂–¥—É –¥–≤—É–º—è –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–∞–º–∏
        print("\n[*] –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É—é —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è –º–µ–∂–¥—É –ø–∞—Ä–∞–º–∏ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤...")
        synced_dfs = self._sync_two_way_matches(synced_dfs)
        
        return synced_dfs
    
    def _sync_three_way_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """–°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ—Ç —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è –≤—Å–µ—Ö —Ç—Ä–µ—Ö –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤"""
        matches = self.comparison_result.get('matches_all_three', [])
        if not matches:
            print("  –ù–µ—Ç —Å–æ–≤–ø–∞–¥–µ–Ω–∏–π –¥–ª—è —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∞—Ü–∏–∏")
            return dfs
        
        total_filled = 0
        skipped_count = 0
        
        for match in matches:
            col_wb = match.get('column_1')
            col_ozon = match.get('column_2')
            col_yandex = match.get('column_3')
            
            if not all([col_wb, col_ozon, col_yandex]):
                continue
            
            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –∏—Å–∫–ª—é—á–µ–Ω–Ω—ã–µ —Å—Ç–æ–ª–±—Ü—ã
            if (is_excluded_column(col_wb) or
                is_excluded_column(col_ozon) or
                is_excluded_column(col_yandex)):
                skipped_count += 1
                continue
            
            # ‚≠ê –î–û–ë–ê–í–¨ –≠–¢–£ –ü–†–û–í–ï–†–ö–£:
            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –≥–∞–±–∞—Ä–∏—Ç—ã - –æ–Ω–∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—é—Ç—Å—è —á–µ—Ä–µ–∑ DimensionsSynchronizer
            # if col_yandex == "–ì–∞–±–∞—Ä–∏—Ç—ã —Å —É–ø–∞–∫–æ–≤–∫–æ–π, —Å–º":
            #     skipped_count += 1
            #     logger.info(f"‚è≠Ô∏è  –ü—Ä–æ–ø—É—â–µ–Ω–æ (–≥–∞–±–∞—Ä–∏—Ç—ã): {col_wb} ‚Üî {col_ozon} ‚Üî {col_yandex}")
            #     continue
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Å—Ç–æ–ª–±—Ü—ã —Å—É—â–µ—Å—Ç–≤—É—é—Ç
            if (col_wb not in dfs['wildberries'].columns or
                col_ozon not in dfs['ozon'].columns or
                col_yandex not in dfs['yandex'].columns):
                continue
            
            # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –º–µ–∂–¥—É —Ç—Ä–µ–º—è —Ñ–∞–π–ª–∞–º–∏
            filled = self._sync_three_columns(
                dfs,
                col_wb, col_ozon, col_yandex
            )
            
            if filled > 0:
                confidence = int(match.get('confidence', 0) * 100)
                print(f"  ‚úì –ó–∞–ø–æ–ª–Ω–µ–Ω–æ {filled} –∑–Ω–∞—á–µ–Ω–∏–π: '{col_wb}' ‚Üî '{col_ozon}' ‚Üî '{col_yandex}' ({confidence}%)")
                total_filled += filled
        
        if skipped_count > 0:
            print(f"[!] –ü—Ä–æ–ø—É—â–µ–Ω–æ {skipped_count} –∏—Å–∫–ª—é—á–µ–Ω–Ω—ã—Ö —Å—Ç–æ–ª–±—Ü–æ–≤")
        print(f"[+] –í—Å–µ–≥–æ –∑–∞–ø–æ–ª–Ω–µ–Ω–æ {total_filled} –ø—É—Å—Ç—ã—Ö —è—á–µ–µ–∫ –≤ —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è—Ö –≤—Å–µ—Ö 3 –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤")
        
        return dfs
    
    def _sync_two_way_matches(self, dfs: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """–°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ—Ç —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è –º–µ–∂–¥—É –ø–∞—Ä–∞–º–∏ –º–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å–æ–≤"""
        
        pairs = [
            ('matches_1_2', 'wildberries', 'ozon', 'column_1', 'column_2'),
            ('matches_1_3', 'wildberries', 'yandex', 'column_1', 'column_3'),
            ('matches_2_3', 'ozon', 'yandex', 'column_2', 'column_3')
        ]
        
        total_filled = 0
        skipped_count = 0
        
        for match_key, mp1, mp2, col_key1, col_key2 in pairs:
            matches = self.comparison_result.get(match_key, [])
            
            if not matches:
                continue
            
            for match in matches:
                col1 = match.get(col_key1)
                col2 = match.get(col_key2)
                
                if not all([col1, col2]):
                    continue
                
                # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –∏—Å–∫–ª—é—á–µ–Ω–Ω—ã–µ —Å—Ç–æ–ª–±—Ü—ã
                if is_excluded_column(col1) or is_excluded_column(col2):
                    skipped_count += 1
                    continue
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Å—Ç–æ–ª–±—Ü—ã —Å—É—â–µ—Å—Ç–≤—É—é—Ç
                if col1 not in dfs[mp1].columns or col2 not in dfs[mp2].columns:
                    continue
                
                # –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –º–µ–∂–¥—É –¥–≤—É–º—è —Ñ–∞–π–ª–∞–º–∏
                filled = self._sync_two_columns(dfs, mp1, mp2, col1, col2)
                
                if filled > 0:
                    confidence = int(match.get('confidence', 0) * 100)
                    print(f"  ‚úì –ó–∞–ø–æ–ª–Ω–µ–Ω–æ {filled} –∑–Ω–∞—á–µ–Ω–∏–π: {mp1}:'{col1}' ‚Üî {mp2}:'{col2}' ({confidence}%)")
                    total_filled += filled
        
        if skipped_count > 0:
            print(f"[!] –ü—Ä–æ–ø—É—â–µ–Ω–æ {skipped_count} –∏—Å–∫–ª—é—á–µ–Ω–Ω—ã—Ö —Å—Ç–æ–ª–±—Ü–æ–≤")
        print(f"[+] –í—Å–µ–≥–æ –∑–∞–ø–æ–ª–Ω–µ–Ω–æ {total_filled} –ø—É—Å—Ç—ã—Ö —è—á–µ–µ–∫ –≤ —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è—Ö –º–µ–∂–¥—É –ø–∞—Ä–∞–º–∏")
        return dfs
    
    def _sync_three_columns(
        self,
        dfs: Dict[str, pd.DataFrame],
        col_wb: str,
        col_ozon: str,
        col_yandex: str
    ) -> int:
        """
        –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ—Ç –¥–∞–Ω–Ω—ã–µ –º–µ–∂–¥—É —Ç—Ä–µ–º—è —Å—Ç–æ–ª–±—Ü–∞–º–∏ –Ω–∞ –æ—Å–Ω–æ–≤–µ –∞—Ä—Ç–∏–∫—É–ª–æ–≤
        Returns:
            –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã—Ö —è—á–µ–µ–∫
        """
        filled_count = 0
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –µ–¥–∏–Ω–∏—Ü—ã –∏–∑–º–µ—Ä–µ–Ω–∏—è –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Å—Ç–æ–ª–±—Ü–∞
        unit_wb = self._detect_unit(col_wb)
        unit_ozon = self._detect_unit(col_ozon)
        unit_yandex = self._detect_unit(col_yandex)
        
        # –°–æ–∑–¥–∞–µ–º —Å–ª–æ–≤–∞—Ä–∏ –¥–ª—è –±—ã—Å—Ç—Ä–æ–≥–æ –ø–æ–∏—Å–∫–∞ –ø–æ –∞—Ä—Ç–∏–∫—É–ª—É
        wb_data = self._create_article_map(dfs['wildberries'], self.article_columns['wildberries'], col_wb)
        ozon_data = self._create_article_map(dfs['ozon'], self.article_columns['ozon'], col_ozon)
        yandex_data = self._create_article_map(dfs['yandex'], self.article_columns['yandex'], col_yandex)
        
        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ —É–Ω–∏–∫–∞–ª—å–Ω—ã–µ –∞—Ä—Ç–∏–∫—É–ª—ã
        all_articles = set(wb_data.keys()) | set(ozon_data.keys()) | set(yandex_data.keys())
        
        for article in all_articles:
            if not article:  # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –ø—É—Å—Ç—ã–µ –∞—Ä—Ç–∏–∫—É–ª—ã
                continue
            
            # –ü–æ–ª—É—á–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è –∏–∑ –≤—Å–µ—Ö —Ç—Ä–µ—Ö –∏—Å—Ç–æ—á–Ω–∏–∫–æ–≤
            values = {
                'wildberries': wb_data.get(article, {}).get('value'),
                'ozon': ozon_data.get(article, {}).get('value'),
                'yandex': yandex_data.get(article, {}).get('value')
            }
            
            # –ù–∞—Ö–æ–¥–∏–º –Ω–µ–ø—É—Å—Ç–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏ –µ–≥–æ –∏—Å—Ç–æ—á–Ω–∏–∫
            source_value = None
            source_unit = None
            for marketplace, val in values.items():
                # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –ø—Ä–æ–≤–µ—Ä—è–µ–º —Ç–∏–ø –∏ –∏–∑–≤–ª–µ–∫–∞–µ–º —Å–∫–∞–ª—è—Ä
                if isinstance(val, pd.Series):
                    if not val.empty:
                        val = val.iloc[0]  # –ë–µ—Ä–µ–º –ø–µ—Ä–≤–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
                    else:
                        val = None
                
                if pd.notna(val) and str(val).strip():
                    source_value = val
                    if marketplace == 'wildberries':
                        source_unit = unit_wb
                    elif marketplace == 'ozon':
                        source_unit = unit_ozon
                    else:
                        source_unit = unit_yandex
                    break
            
            if source_value is None:
                continue
            
            # WB
            if article in wb_data:
                val_wb = values['wildberries']
                if isinstance(val_wb, pd.Series):
                    val_wb = val_wb.iloc[0] if not val_wb.empty else None
                
                if pd.isna(val_wb) or not str(val_wb).strip():
                    idx = wb_data[article]['index']
                    series = dfs['wildberries'][col_wb]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]  # –ë–µ—Ä–µ–º –ø–µ—Ä–≤—ã–π —Å—Ç–æ–ª–±–µ—Ü –µ—Å–ª–∏ —ç—Ç–æ DataFrame
                    col_dtype = series.dtype
                    converted_value = self._convert_value(source_value, source_unit, unit_wb)
                    
                    # –ü—Ä–æ–≤–µ—Ä–∫–∞ validation —á–µ—Ä–µ–∑ AI
                    final_value = self._validate_with_ai(converted_value, 'wildberries', col_wb)
                    
                    try:
                        # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get('wildberries', {}).get(col_wb):
                            # –ù–µ—Ç validation - –∑–∞–ø–∏—Å—ã–≤–∞–µ–º –∫–∞–∫ –µ—Å—Ç—å
                            value_to_set = converted_value
                        else:
                            # –ï—Å—Ç—å validation –Ω–æ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ - –ù–ï –∑–∞–ø–∏—Å—ã–≤–∞–µ–º
                            logger.warning(f"‚ö†Ô∏è [WB] –ü—Ä–æ–ø—É—â–µ–Ω–æ '{converted_value}' –¥–ª—è '{col_wb}' (–Ω–µ –ø—Ä–æ—à–ª–æ validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs['wildberries'].at[idx, col_wb] = value_to_set
                        filled_count += 1
                        self._log_change('wildberries', article, col_wb, value_to_set, source_marketplace='ozon' if source_unit == unit_ozon else ('yandex' if source_unit == unit_yandex else 'wildberries'))
                    except Exception:
                        pass
            
            # OZON
            if article in ozon_data:
                val_ozon = values['ozon']
                if isinstance(val_ozon, pd.Series):
                    val_ozon = val_ozon.iloc[0] if not val_ozon.empty else None
                
                if pd.isna(val_ozon) or not str(val_ozon).strip():
                    idx = ozon_data[article]['index']
                    series = dfs['ozon'][col_ozon]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype
                    converted_value = self._convert_value(source_value, source_unit, unit_ozon)
                    
                    final_value = self._validate_with_ai(converted_value, 'ozon', col_ozon)
                    
                    try:
                        # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get('ozon', {}).get(col_ozon):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"‚ö†Ô∏è [OZON] –ü—Ä–æ–ø—É—â–µ–Ω–æ '{converted_value}' –¥–ª—è '{col_ozon}' (–Ω–µ –ø—Ä–æ—à–ª–æ validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs['ozon'].at[idx, col_ozon] = value_to_set
                        filled_count += 1
                        self._log_change('ozon', article, col_ozon, value_to_set, source_marketplace='wildberries' if source_unit == unit_wb else ('yandex' if source_unit == unit_yandex else 'ozon'))
                    except Exception:
                        pass
            
            # YANDEX
            if article in yandex_data:
                val_yandex = values['yandex']
                if isinstance(val_yandex, pd.Series):
                    val_yandex = val_yandex.iloc[0] if not val_yandex.empty else None
                
                if pd.isna(val_yandex) or not str(val_yandex).strip():
                    idx = yandex_data[article]['index']
                    series = dfs['yandex'][col_yandex]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype
                    
                    # ‚úÖ –ü–†–û–í–ï–†–ö–ê: –≠—Ç–æ –∫–æ–º–ø–æ–∑–∏—Ç–Ω—ã–π —Å—Ç–æ–ª–±–µ—Ü –≥–∞–±–∞—Ä–∏—Ç–æ–≤?
                    if col_yandex == "–ì–∞–±–∞—Ä–∏—Ç—ã —Å —É–ø–∞–∫–æ–≤–∫–æ–π, —Å–º":
                        # –§–æ—Ä–º–∏—Ä—É–µ–º –∫–æ–º–ø–æ–∑–∏—Ç–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏–∑ —Ç—Ä—ë—Ö –∏–∑–º–µ—Ä–µ–Ω–∏–π
                        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∏—Å—Ç–æ—á–Ω–∏–∫ (WB –∏–ª–∏ Ozon)
                        if source_unit == unit_wb and article in wb_data:
                            # –ò—Å—Ç–æ—á–Ω–∏–∫ - WB
                            wb_row = dfs['wildberries'][dfs['wildberries'][self.article_columns['wildberries']].astype(str).str.strip() == article].iloc[0]
                            length = wb_row.get('–î–ª–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)', None)
                            width = wb_row.get('–®–∏—Ä–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)', None)
                            height = wb_row.get('–í—ã—Å–æ—Ç–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)', None)
                            
                            if all(pd.notna(v) for v in [length, width, height]):
                                composite = DimensionsSynchronizer.format_composite_dimensions(
                                    float(length), float(width), float(height)
                                )
                                dfs['yandex'].at[idx, col_yandex] = composite
                                filled_count += 1
                                self._log_change('yandex', article, col_yandex, composite, source_marketplace='wildberries')
                                continue
                                
                        elif source_unit == unit_ozon and article in ozon_data:
                            # –ò—Å—Ç–æ—á–Ω–∏–∫ - Ozon (–∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –º–º ‚Üí —Å–º)
                            ozon_row = dfs['ozon'][dfs['ozon'][self.article_columns['ozon']].astype(str).str.strip() == article].iloc[0]
                            length_mm = ozon_row.get('–î–ª–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏, –º–º*', None)
                            width_mm = ozon_row.get('–®–∏—Ä–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏, –º–º*', None)
                            height_mm = ozon_row.get('–í—ã—Å–æ—Ç–∞ —É–ø–∞–∫–æ–≤–∫–∏, –º–º*', None)
                            
                            if all(pd.notna(v) for v in [length_mm, width_mm, height_mm]):
                                composite = DimensionsSynchronizer.format_composite_dimensions(
                                    DimensionsSynchronizer.mm_to_cm(float(length_mm)),
                                    DimensionsSynchronizer.mm_to_cm(float(width_mm)),
                                    DimensionsSynchronizer.mm_to_cm(float(height_mm))
                                )
                                dfs['yandex'].at[idx, col_yandex] = composite
                                filled_count += 1
                                self._log_change('yandex', article, col_yandex, composite, source_marketplace='ozon')
                                continue
                    
                    # –û–±—ã—á–Ω–∞—è –ª–æ–≥–∏–∫–∞ –¥–ª—è –æ—Å—Ç–∞–ª—å–Ω—ã—Ö —Å—Ç–æ–ª–±—Ü–æ–≤
                    converted_value = self._convert_value(source_value, source_unit, unit_yandex)
                    final_value = self._validate_with_ai(converted_value, 'yandex', col_yandex)
                    
                    try:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get('yandex', {}).get(col_yandex):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"‚ö†Ô∏è [YANDEX] –ü—Ä–æ–ø—É—â–µ–Ω–æ '{converted_value}' –¥–ª—è '{col_yandex}' (–Ω–µ –ø—Ä–æ—à–ª–æ validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        
                        dfs['yandex'].at[idx, col_yandex] = value_to_set
                        filled_count += 1
                        self._log_change('yandex', article, col_yandex, value_to_set, 
                                    source_marketplace='wildberries' if source_unit == unit_wb else ('ozon' if source_unit == unit_ozon else 'yandex'))
                    except Exception:
                        pass
        
        return filled_count
    
    def _validate_with_ai(self, value, marketplace: str, column_name: str) -> Optional[str]:
        """
        –ü—Ä–æ–≤–µ—Ä—è–µ—Ç –∑–Ω–∞—á–µ–Ω–∏–µ —á–µ—Ä–µ–∑ AI –µ—Å–ª–∏ –µ—Å—Ç—å validation
        Returns:
            –°–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏–ª–∏ None –µ—Å–ª–∏ –Ω–µ—Ç validation
        """
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –µ—Å—Ç—å –ª–∏ validation –¥–ª—è —ç—Ç–æ–≥–æ —Å—Ç–æ–ª–±—Ü–∞
        allowed_values = self.column_validations.get(marketplace, {}).get(column_name)
        
        if not allowed_values or not self.ai_comparator:
            return None
        
        value_str = str(value).strip()
        
        # –§—É–Ω–∫—Ü–∏—è –Ω–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏–∏
        def normalize(text: str) -> str:
            """–ù–æ—Ä–º–∞–ª–∏–∑—É–µ—Ç —Ç–µ–∫—Å—Ç: –Ω–∏–∂–Ω–∏–π —Ä–µ–≥–∏—Å—Ç—Ä, —ë‚Üí–µ"""
            return text.lower().replace('—ë', '–µ').strip()
        
        # –§—É–Ω–∫—Ü–∏—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è —á–∏—Å–ª–∞
        def extract_number(text: str) -> Optional[str]:
            """–ò–∑–≤–ª–µ–∫–∞–µ—Ç –ø–µ—Ä–≤–æ–µ —á–∏—Å–ª–æ –∏–∑ —Å—Ç—Ä–æ–∫–∏ —Ç–∏–ø–∞ '1 —à—Ç', '2 –∫–æ–º–ø—Ä–µ—Å—Å–æ—Ä–∞'"""
            import re
            numbers = re.findall(r'\d+', text)
            return numbers[0] if numbers else None
        
        # –§—É–Ω–∫—Ü–∏—è –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è
        def log_match(original: str, matched: str, method: str):
            """–ó–∞–ø–∏—Å—ã–≤–∞–µ—Ç —É—Å–ø–µ—à–Ω–æ–µ —Å–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–∏–µ –≤ –ª–æ–≥"""
            self.ai_validation_log.append({
                '–ú–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å': marketplace.upper(),
                '–°—Ç–æ–ª–±–µ—Ü': column_name,
                '–ò—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ': original,
                '–°–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–æ —Å': matched,
                '–ú–µ—Ç–æ–¥': method
            })
        
        # 1. –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ
        if value_str in allowed_values:
            logger.info(f"[_validate_with_ai] –¢–û–ß–ù–û–ï —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ: '{value_str}'")
            log_match(value_str, value_str, '–¢–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ')
            return value_str
        
        # 2. –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å –Ω–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏–µ–π (—Ä–µ–≥–∏—Å—Ç—Ä + —ë/–µ)
        value_normalized = normalize(value_str)
        for allowed in allowed_values:
            if normalize(allowed) == value_normalized:
                logger.info(f"[_validate_with_ai] –°–æ–≤–ø–∞–¥–µ–Ω–∏–µ —Å –Ω–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏–µ–π: '{value_str}' ‚Üí '{allowed}'")
                log_match(value_str, allowed, '–ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è (—Ä–µ–≥–∏—Å—Ç—Ä/—ë-–µ)')
                return allowed
        
        # 3. –ò–∑–≤–ª–µ–∫–∞–µ–º —á–∏—Å–ª–æ –µ—Å–ª–∏ —ç—Ç–æ —á–∏—Å–ª–æ–≤–æ–µ –ø–æ–ª–µ
        number = extract_number(value_str)
        if number:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ —á–∏—Å–ª–∞
            if number in allowed_values:
                logger.info(f"[_validate_with_ai] –ò–∑–≤–ª–µ—á–µ–Ω–æ —á–∏—Å–ª–æ: '{value_str}' ‚Üí '{number}'")
                log_match(value_str, number, '–ò–∑–≤–ª–µ—á–µ–Ω–∏–µ —á–∏—Å–ª–∞')
                return number
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å –Ω–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏–µ–π
            for allowed in allowed_values:
                if extract_number(allowed) == number:
                    logger.info(f"[_validate_with_ai] –°–æ–≤–ø–∞–¥–µ–Ω–∏–µ –ø–æ —á–∏—Å–ª—É: '{value_str}' ‚Üí '{allowed}'")
                    log_match(value_str, allowed, '–ò–∑–≤–ª–µ—á–µ–Ω–∏–µ —á–∏—Å–ª–∞')
                    return allowed
        
        # 4. –ü—Ä–æ–≤–µ—Ä—è–µ–º —á–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ (–ø–æ —Å–ª–æ–≤–∞–º)
        value_words = set(value_normalized.split())
        for allowed in allowed_values:
            allowed_words = set(normalize(allowed).split())
            
            # –ï—Å–ª–∏ –≤—Å–µ —Å–ª–æ–≤–∞ –∏–∑ value –µ—Å—Ç—å –≤ allowed
            if value_words and value_words.issubset(allowed_words):
                logger.info(f"[_validate_with_ai] –ß–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ: '{value_str}' ‚Üí '{allowed}'")
                log_match(value_str, allowed, '–ß–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ (—Å–ª–æ–≤–∞)')
                return allowed
        
        # 5. –°–ø—Ä–∞—à–∏–≤–∞–µ–º AI (–±–µ–∑ –∫—ç—à–∞!)
        logger.info(f"ü§ñ [AI] –ü—Ä–æ–≤–µ—Ä—è—é '{value_str}' –¥–ª—è —Å—Ç–æ–ª–±—Ü–∞ '{column_name}'...")
        matched_value = self.ai_comparator.match_value_with_list(value_str, allowed_values, column_name=column_name)  # ‚Üê –î–û–ë–ê–í–ò–¢–¨!)
        
        if matched_value:
            logger.info(f"‚úÖ [AI] –ù–∞–π–¥–µ–Ω–æ: '{value_str}' ‚Üí '{matched_value}'")
            log_match(value_str, matched_value, 'AI –∑–∞–ø—Ä–æ—Å')
            return matched_value
        else:
            logger.warning(f"‚ùå [AI] –ù–µ –Ω–∞–π–¥–µ–Ω–æ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –¥–ª—è '{value_str}'")
            return None

    
    def _sync_two_columns(
        self,
        dfs: Dict[str, pd.DataFrame],
        mp1: str,
        mp2: str,
        col1: str,
        col2: str
    ) -> int:
        """
        –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä—É–µ—Ç –¥–∞–Ω–Ω—ã–µ –º–µ–∂–¥—É –¥–≤—É–º—è —Å—Ç–æ–ª–±—Ü–∞–º–∏ –Ω–∞ –æ—Å–Ω–æ–≤–µ –∞—Ä—Ç–∏–∫—É–ª–æ–≤
        Returns:
            –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –∑–∞–ø–æ–ª–Ω–µ–Ω–Ω—ã—Ö —è—á–µ–µ–∫
        """
        filled_count = 0
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –µ–¥–∏–Ω–∏—Ü—ã –∏–∑–º–µ—Ä–µ–Ω–∏—è
        unit1 = self._detect_unit(col1)
        unit2 = self._detect_unit(col2)
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Å—Ç–æ–ª–±—Ü—ã –∞—Ä—Ç–∏–∫—É–ª–æ–≤
        article_col1 = self.article_columns[mp1]
        article_col2 = self.article_columns[mp2]
        
        # –°–æ–∑–¥–∞–µ–º —Å–ª–æ–≤–∞—Ä–∏ –¥–ª—è –±—ã—Å—Ç—Ä–æ–≥–æ –ø–æ–∏—Å–∫–∞
        data1 = self._create_article_map(dfs[mp1], article_col1, col1)
        data2 = self._create_article_map(dfs[mp2], article_col2, col2)
        
        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ —É–Ω–∏–∫–∞–ª—å–Ω—ã–µ –∞—Ä—Ç–∏–∫—É–ª—ã
        all_articles = set(data1.keys()) | set(data2.keys())
        
        for article in all_articles:
            if not article:
                continue
            
            # –ü–æ–ª—É—á–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è
            val1 = data1.get(article, {}).get('value')
            val2 = data2.get(article, {}).get('value')
            
            # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –ø—Ä–æ–≤–µ—Ä—è–µ–º Series
            if isinstance(val1, pd.Series):
                val1 = val1.iloc[0] if not val1.empty else None
            if isinstance(val2, pd.Series):
                val2 = val2.iloc[0] if not val2.empty else None
            
            # –ó–∞–ø–æ–ª–Ω—è–µ–º –ø—É—Å—Ç—ã–µ —è—á–µ–π–∫–∏
            if article in data1 and article in data2:
                # –ï—Å–ª–∏ –≤ –ø–µ—Ä–≤–æ–º –ø—É—Å—Ç–æ, –∞ –≤–æ –≤—Ç–æ—Ä–æ–º –µ—Å—Ç—å
                if (pd.isna(val1) or not str(val1).strip()) and pd.notna(val2) and str(val2).strip():
                    idx = data1[article]['index']
                    series = dfs[mp1][col1]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype  # ‚úÖ
                    converted_value = self._convert_value(val2, unit2, unit1)
                    
                    final_value = self._validate_with_ai(converted_value, mp1, col1)
                    
                    try:
                        # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get(mp1, {}).get(col1):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"‚ö†Ô∏è [{mp1.upper()}] –ü—Ä–æ–ø—É—â–µ–Ω–æ '{converted_value}' –¥–ª—è '{col1}' (–Ω–µ –ø—Ä–æ—à–ª–æ validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs[mp1].at[idx, col1] = value_to_set
                        filled_count += 1
                        self._log_change(mp1, article, col1, value_to_set, source_marketplace=mp2)
                    except Exception:
                        pass
                
                # –ï—Å–ª–∏ –≤–æ –≤—Ç–æ—Ä–æ–º –ø—É—Å—Ç–æ, –∞ –≤ –ø–µ—Ä–≤–æ–º –µ—Å—Ç—å
                elif (pd.isna(val2) or not str(val2).strip()) and pd.notna(val1) and str(val1).strip():
                    idx = data2[article]['index']
                    series = dfs[mp2][col2]
                    if isinstance(series, pd.DataFrame):
                        series = series.iloc[:, 0]
                    col_dtype = series.dtype  # ‚úÖ
                    converted_value = self._convert_value(val1, unit1, unit2)
                    
                    final_value = self._validate_with_ai(converted_value, mp2, col2)
                    
                    try:
                        # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï:
                        if final_value:
                            value_to_set = final_value
                        elif not self.column_validations.get(mp2, {}).get(col2):
                            value_to_set = converted_value
                        else:
                            logger.warning(f"‚ö†Ô∏è [{mp2.upper()}] –ü—Ä–æ–ø—É—â–µ–Ω–æ '{converted_value}' –¥–ª—è '{col2}' (–Ω–µ –ø—Ä–æ—à–ª–æ validation)")
                            continue
                        
                        if pd.api.types.is_numeric_dtype(col_dtype):
                            value_to_set = pd.to_numeric(value_to_set, errors='coerce')
                        dfs[mp2].at[idx, col2] = value_to_set
                        filled_count += 1
                        self._log_change(mp2, article, col2, value_to_set, source_marketplace=mp1)
                    except Exception:
                        pass
        
        return filled_count
    
    def _create_article_map(self, df: pd.DataFrame, article_col: str, value_col: str) -> Dict:
        """
        –°–æ–∑–¥–∞–µ—Ç —Å–ª–æ–≤–∞—Ä—å –¥–ª—è –±—ã—Å—Ç—Ä–æ–≥–æ –ø–æ–∏—Å–∫–∞ –∑–Ω–∞—á–µ–Ω–∏–π –ø–æ –∞—Ä—Ç–∏–∫—É–ª—É
        
        Returns:
            –°–ª–æ–≤–∞—Ä—å {–∞—Ä—Ç–∏–∫—É–ª: {'value': –∑–Ω–∞—á–µ–Ω–∏–µ, 'index': –∏–Ω–¥–µ–∫—Å —Å—Ç—Ä–æ–∫–∏}}
        """
        article_map = {}
        
        if article_col not in df.columns or value_col not in df.columns:
            return article_map
        
        for idx, row in df.iterrows():
            article = row.get(article_col)
            value = row.get(value_col)
            
            if pd.notna(article):
                article_str = str(article).strip()
                if article_str:
                    article_map[article_str] = {
                        'value': value,
                        'index': idx
                    }
        
        return article_map
    
    def _log_change(self, marketplace: str, article: str, column: str, new_value, source_marketplace: str = None):
        """–õ–æ–≥–∏—Ä—É–µ—Ç –ø—Ä–æ–∏–∑–≤–µ–¥–µ–Ω–Ω–æ–µ –∏–∑–º–µ–Ω–µ–Ω–∏–µ"""
        self.changes_log[marketplace].append({
            'article': article,
            'column': column,
            'new_value': str(new_value),
            'source_marketplace': source_marketplace  # ‚Üê –î–û–ë–ê–í–õ–ï–ù–û
        })
    
    def _postprocess_wb_dimensions(self, dfs: Dict[str, pd.DataFrame]) -> int:
        """
        –ü–æ—Å—Ç–æ–±—Ä–∞–±–æ—Ç–∫–∞ –≥–∞–±–∞—Ä–∏—Ç–æ–≤ WB: –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è –∏–∑ –º–º –≤ —Å–º –µ—Å–ª–∏ –¥–∞–Ω–Ω—ã–µ –ø—Ä–∏—à–ª–∏ –∏–∑ Ozon
        
        Returns:
            –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å–∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
        """
        if 'wildberries' not in dfs:
            return 0
        
        converted_count = 0
        df_wb = dfs['wildberries']
        
        # –°—Ç–æ–ª–±—Ü—ã –≥–∞–±–∞—Ä–∏—Ç–æ–≤ WB
        wb_dimension_columns = [
            '–î–ª–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)',
            '–®–∏—Ä–∏–Ω–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)',
            '–í—ã—Å–æ—Ç–∞ —É–ø–∞–∫–æ–≤–∫–∏ (—Ü–µ–ª–æ–µ —á–∏—Å–ª–æ)'
        ]
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ —Å—Ç–æ–ª–±—Ü–æ–≤
        for col_name in wb_dimension_columns:
            if col_name not in df_wb.columns:
                return 0
        
        # –ü—Ä–æ—Ö–æ–¥–∏–º –ø–æ –≤—Å–µ–º –∏–∑–º–µ–Ω–µ–Ω–∏—è–º WB –∏–∑ Ozon
        if 'wildberries' not in self.changes_log:
            return 0
        
        for change in self.changes_log['wildberries']:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º: —ç—Ç–æ –≥–∞–±–∞—Ä–∏—Ç –ò –∏—Å—Ç–æ—á–Ω–∏–∫ - Ozon
            if (change.get('source_marketplace') == 'ozon' and 
                change.get('column') in wb_dimension_columns):
                
                article = change.get('article')
                column = change.get('column')
                
                # –ù–∞—Ö–æ–¥–∏–º —Å—Ç—Ä–æ–∫—É —Å —ç—Ç–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º
                mask = df_wb[self.article_columns['wildberries']].astype(str).str.strip() == str(article).strip()
                if mask.any():
                    idx = df_wb[mask].index[0]
                    value = df_wb.at[idx, column]
                    
                    # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –µ—Å–ª–∏ >= 100 (–∑–∞—â–∏—Ç–∞ –æ—Ç –ø–æ–≤—Ç–æ—Ä–Ω–æ–π –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏–∏)
                    if pd.notna(value):
                        try:
                            numeric_value = float(value)
                            if numeric_value >= 100:
                                converted_value = round(numeric_value / 10, 1)
                                df_wb.at[idx, column] = converted_value
                                converted_count += 1
                                logger.info(f"  ‚úì [{article}] {column}: {numeric_value} –º–º ‚Üí {converted_value} —Å–º")
                        except (ValueError, TypeError):
                            pass
        
        return converted_count

    
    def _get_validation_list_values(self, ws, row_idx: int, col_idx: int) -> List[str]:
        """
        –ü–æ–ª—É—á–∞–µ—Ç —Å–ø–∏—Å–æ–∫ –¥–æ–ø—É—Å—Ç–∏–º—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π –∏–∑ data validation —è—á–µ–π–∫–∏
        
        Returns:
            –°–ø–∏—Å–æ–∫ –¥–æ–ø—É—Å—Ç–∏–º—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π –∏–ª–∏ –ø—É—Å—Ç–æ–π —Å–ø–∏—Å–æ–∫
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        
        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
        
        # –ü—Ä–æ—Ö–æ–¥–∏–º –ø–æ –≤—Å–µ–º data validation –ø—Ä–∞–≤–∏–ª–∞–º
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list":
                continue
            
            # –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –ø—Ä–æ–≤–µ—Ä—è–µ–º –ø—Ä–∏–Ω–∞–¥–ª–µ–∂–Ω–æ—Å—Ç—å —è—á–µ–π–∫–∏ –∫ –¥–∏–∞–ø–∞–∑–æ–Ω—É validation
            # –ò—Å–ø–æ–ª—å–∑—É–µ–º –ø—Ä–∞–≤–∏–ª—å–Ω—É—é –ø—Ä–æ–≤–µ—Ä–∫—É —á–µ—Ä–µ–∑ sqref (string reference)
            if dv.sqref is None:
                continue
                
            # sqref –º–æ–∂–µ—Ç —Å–æ–¥–µ—Ä–∂–∞—Ç—å –Ω–µ—Å–∫–æ–ª—å–∫–æ –¥–∏–∞–ø–∞–∑–æ–Ω–æ–≤, —Ä–∞–∑–¥–µ–ª–µ–Ω–Ω—ã—Ö –ø—Ä–æ–±–µ–ª–∞–º–∏
            # –ù–∞–ø—Ä–∏–º–µ—Ä: "B2:B100 D2:D100"
            ranges = str(dv.sqref).split()
            
            cell_in_range = False
            for range_str in ranges:
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤—Ö–æ–¥–∏—Ç –ª–∏ –Ω–∞—à–∞ —è—á–µ–π–∫–∞ –≤ –¥–∏–∞–ø–∞–∑–æ–Ω
                if ':' in range_str:
                    # –î–∏–∞–ø–∞–∑–æ–Ω —Ç–∏–ø–∞ A2:A100
                    try:
                        from openpyxl.utils import range_boundaries
                        min_col, min_row, max_col, max_row = range_boundaries(range_str)
                        
                        if (min_col <= col_idx <= max_col and 
                            min_row <= row_idx <= max_row):
                            cell_in_range = True
                            break
                    except:
                        pass
                else:
                    # –û–¥–∏–Ω–æ—á–Ω–∞—è —è—á–µ–π–∫–∞ —Ç–∏–ø–∞ A2
                    if range_str == cell_ref:
                        cell_in_range = True
                        break
            
            if not cell_in_range:
                continue
            
            # –ù–∞—à–ª–∏ validation –¥–ª—è —ç—Ç–æ–π —è—á–µ–π–∫–∏, –∏–∑–≤–ª–µ–∫–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è
            if dv.formula1:
                formula = dv.formula1
                
                # –°–ø–∏—Å–æ–∫ –∑–∞–¥–∞–Ω –ø—Ä—è–º–æ: "–ö—Ä–∞—Å–Ω—ã–π,–°–∏–Ω–∏–π,–ó–µ–ª–µ–Ω—ã–π"
                if formula.startswith('"') and formula.endswith('"'):
                    values = formula.strip('"').split(',')
                    return [v.strip() for v in values]
                
                # –°–ø–∏—Å–æ–∫ –∑–∞–¥–∞–Ω —á–µ—Ä–µ–∑ –¥–∏–∞–ø–∞–∑–æ–Ω: $A$1:$A$10 –∏–ª–∏ Sheet1!$A$1:$A$10
                elif ':' in formula:
                    try:
                        # –£–±–∏—Ä–∞–µ–º $ –¥–ª—è –ø–∞—Ä—Å–∏–Ω–≥–∞
                        clean_formula = formula.replace('$', '')
                        
                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —Å—Å—ã–ª–∫–∞ –Ω–∞ –¥—Ä—É–≥–æ–π –ª–∏—Å—Ç
                        if '!' in clean_formula:
                            sheet_name, range_ref = clean_formula.split('!')
                            target_ws = ws.parent[sheet_name]
                        else:
                            range_ref = clean_formula
                            target_ws = ws
                        
                        # –ò–∑–≤–ª–µ–∫–∞–µ–º –∑–Ω–∞—á–µ–Ω–∏—è –∏–∑ –¥–∏–∞–ø–∞–∑–æ–Ω–∞
                        values = []
                        for row in target_ws[range_ref]:
                            for cell in row:
                                if cell.value is not None:
                                    values.append(str(cell.value).strip())
                        
                        return values
                        
                    except Exception as e:
                        print(f"      [!] –ù–µ —É–¥–∞–ª–æ—Å—å –∏–∑–≤–ª–µ—á—å —Å–ø–∏—Å–æ–∫ –∏–∑ –¥–∏–∞–ø–∞–∑–æ–Ω–∞ {formula}: {e}")
                        return []
        
        return []
    
    def _create_ai_log_sheet(self, output_paths: Dict[str, str]):
        """–°–æ–∑–¥–∞—ë—Ç –ª–∏—Å—Ç —Å –ª–æ–≥–∞–º–∏ AI-—Å–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–∏–π"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # –°–æ–∑–¥–∞—ë–º DataFrame –∏–∑ –ª–æ–≥–æ–≤
        df_log = pd.DataFrame(self.ai_validation_log)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –ª–∏—Å—Ç –≤ –∫–∞–∂–¥—ã–π –≤—ã—Ö–æ–¥–Ω–æ–π —Ñ–∞–π–ª
        for marketplace, output_path in output_paths.items():
            try:
                # –ó–∞–≥—Ä—É–∂–∞–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π —Ñ–∞–π–ª
                wb = load_workbook(output_path)
                
                # –°–æ–∑–¥–∞—ë–º –Ω–æ–≤—ã–π –ª–∏—Å—Ç
                if 'AI_–õ–æ–≥–∏' in wb.sheetnames:
                    del wb['AI_–õ–æ–≥–∏']
                ws = wb.create_sheet('AI_–õ–æ–≥–∏', 0)  # 0 = –ø–µ—Ä–≤—ã–π –ª–∏—Å—Ç
                
                # –ó–∞–≥–æ–ª–æ–≤–∫–∏
                headers = ['–ú–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å', '–°—Ç–æ–ª–±–µ—Ü', '–ò—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ', '–°–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–æ —Å', '–ú–µ—Ç–æ–¥']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # –î–∞–Ω–Ω—ã–µ
                for row_idx, row_data in enumerate(self.ai_validation_log, start=2):
                    ws.cell(row=row_idx, column=1, value=row_data['–ú–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å'])
                    ws.cell(row=row_idx, column=2, value=row_data['–°—Ç–æ–ª–±–µ—Ü'])
                    ws.cell(row=row_idx, column=3, value=row_data['–ò—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ'])
                    ws.cell(row=row_idx, column=4, value=row_data['–°–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–æ —Å'])
                    ws.cell(row=row_idx, column=5, value=row_data['–ú–µ—Ç–æ–¥'])
                
                # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                # –°–æ—Ö—Ä–∞–Ω—è–µ–º
                wb.save(output_path)
                logger.info(f"‚úÖ –õ–∏—Å—Ç 'AI_–õ–æ–≥–∏' –¥–æ–±–∞–≤–ª–µ–Ω –≤ {output_path}")
                
            except Exception as e:
                logger.error(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ AI-–ª–æ–≥–∞ –¥–ª—è {marketplace}: {e}")
    
    def _create_ai_log_sheet_in_report(self, report_path: str):
        """–°–æ–∑–¥–∞—ë—Ç –ª–∏—Å—Ç —Å –ª–æ–≥–∞–º–∏ AI-—Å–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–∏–π –≤ —Ñ–∞–π–ª–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞"""
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        try:
            # –ó–∞–≥—Ä—É–∂–∞–µ–º —Ñ–∞–π–ª —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞
            wb = load_workbook(report_path)
            
            # –£–¥–∞–ª—è–µ–º —Å—Ç–∞—Ä—ã–π –ª–∏—Å—Ç –µ—Å–ª–∏ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
            if 'AI_–õ–æ–≥–∏' in wb.sheetnames:
                del wb['AI_–õ–æ–≥–∏']
            
            # –°–æ–∑–¥–∞—ë–º –Ω–æ–≤—ã–π –ª–∏—Å—Ç (–ø–µ—Ä–≤—ã–º –ø–æ—Å–ª–µ –≥–ª–∞–≤–Ω–æ–≥–æ)
            ws = wb.create_sheet('AI_–õ–æ–≥–∏', 1)  # –ò–Ω–¥–µ–∫—Å 1 = –≤—Ç–æ—Ä–æ–π –ª–∏—Å—Ç
            
            # –ó–∞–≥–æ–ª–æ–≤–∫–∏
            headers = ['–ú–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å', '–°—Ç–æ–ª–±–µ—Ü', '–ò—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ', '–°–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–æ —Å', '–ú–µ—Ç–æ–¥']
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # –î–∞–Ω–Ω—ã–µ
            for row_idx, row_data in enumerate(self.ai_validation_log, start=2):
                ws.cell(row=row_idx, column=1, value=row_data.get('–ú–∞—Ä–∫–µ—Ç–ø–ª–µ–π—Å', ''))
                ws.cell(row=row_idx, column=2, value=row_data.get('–°—Ç–æ–ª–±–µ—Ü', ''))
                ws.cell(row=row_idx, column=3, value=row_data.get('–ò—Å—Ö–æ–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ', ''))
                ws.cell(row=row_idx, column=4, value=row_data.get('–°–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–æ —Å', ''))
                ws.cell(row=row_idx, column=5, value=row_data.get('–ú–µ—Ç–æ–¥', ''))
            
            # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                
                for cell in col:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # –ú–æ—Ä–æ–∑–∏–º —à–∞–ø–∫—É
            ws.freeze_panes = 'A2'
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º
            wb.save(report_path)
            logger.info(f"‚úÖ –õ–∏—Å—Ç 'AI_–õ–æ–≥–∏' –¥–æ–±–∞–≤–ª–µ–Ω –≤ {report_path}")
            logger.info(f"üìä –í—Å–µ–≥–æ –∑–∞–ø–∏—Å–µ–π AI-–ª–æ–≥–æ–≤: {len(self.ai_validation_log)}")
            
        except Exception as e:
            logger.error(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ AI-–ª–æ–≥–∞: {e}")

    
    def _save_results(self, dfs: Dict[str, pd.DataFrame], output_paths: Dict[str, str]):
        """–°–æ—Ö—Ä–∞–Ω—è–µ—Ç —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ –≤ —Ñ–∞–π–ª—ã"""
        print("\n[*] –°–æ—Ö—Ä–∞–Ω—è—é —Å–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ...")
        
        stats = {'saved': 0, 'ai_matched': 0, 'validation_conflicts': 0, 'skipped': 0}
        
        for marketplace, df in dfs.items():
            output_path = output_paths.get(marketplace)
            if not output_path:
                continue
            
            config = FILE_CONFIGS[marketplace]
            original_file = self.original_file_paths[marketplace]
            
            print(f"\n[*] –û–±—Ä–∞–±–æ—Ç–∫–∞ {config['display_name']}...")
            
            # –°–±—Ä–∞—Å—ã–≤–∞–µ–º –∏–Ω–¥–µ–∫—Å—ã –ü–ï–†–ï–î —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ–º
            df = df.reset_index(drop=True)
            
            # –û—Ç–∫—Ä—ã–≤–∞–µ–º –û–†–ò–ì–ò–ù–ê–õ–¨–ù–´–ô —Ñ–∞–π–ª —á–µ—Ä–µ–∑ openpyxl
            wb = load_workbook(original_file)
            ws = wb[config['sheet_name']]
            
            header_row = config['header_row']
            data_start_row = config.get('data_start_row', header_row + 1)
            
            # üÜï –°–û–ó–î–ê–Å–ú –ú–ê–ü–ü–ò–ù–ì: DataFrame —Å—Ç–æ–ª–±–µ—Ü ‚Üí Excel —Å—Ç–æ–ª–±—Ü—ã (—Å —É—á—ë—Ç–æ–º –¥—É–±–ª–∏–∫–∞—Ç–æ–≤!)
            df_to_excel_mapping = {}  # {df_col_name: [excel_col_1, excel_col_2, ...]}
            header_count = {}
            
            # –ß–∏—Ç–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏ –∏–∑ Excel
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value:
                    col_name = str(cell.value).strip()
                    
                    # –ü–µ—Ä–≤–æ–µ –≤—Ö–æ–∂–¥–µ–Ω–∏–µ - –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω–æ–µ –∏–º—è
                    if col_name not in header_count:
                        header_count[col_name] = 0
                        df_col = col_name  # "–í–µ—Å —Å —É–ø–∞–∫–æ–≤–∫–æ–π (–∫–≥)"
                    else:
                        # –î—É–±–ª–∏–∫–∞—Ç - —Å —Å—É—Ñ—Ñ–∏–∫—Å–æ–º
                        header_count[col_name] += 1
                        df_col = f"{col_name}{header_count[col_name]}"  # "–í–µ—Å —Å —É–ø–∞–∫–æ–≤–∫–æ–π (–∫–≥)1"
                    
                    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –º–∞–ø–ø–∏–Ω–≥
                    if df_col not in df_to_excel_mapping:
                        df_to_excel_mapping[df_col] = []
                    df_to_excel_mapping[df_col].append(col_idx)
            
            # üÜï –°–ò–ù–•–†–û–ù–ò–ó–ê–¶–ò–Ø –î–£–ë–õ–ò–ö–ê–¢–û–í –ü–ï–†–ï–î –ó–ê–ü–ò–°–¨–Æ
            if marketplace in self.original_column_names:
                renamed_map = self.original_column_names[marketplace]['renamed']
                
                for duplicated_name, original_name in renamed_map.items():
                    # duplicated_name = "–í–µ—Å —Å —É–ø–∞–∫–æ–≤–∫–æ–π (–∫–≥)1"
                    # original_name = "–í–µ—Å —Å —É–ø–∞–∫–æ–≤–∫–æ–π (–∫–≥)"
                    
                    if original_name in df.columns and duplicated_name in df.columns:
                        # –ö–æ–ø–∏—Ä—É–µ–º –∑–Ω–∞—á–µ–Ω–∏—è –∏–∑ –æ—Ä–∏–≥–∏–Ω–∞–ª–∞ –≤ –¥—É–±–ª–∏–∫–∞—Ç
                        for idx in df.index:
                            original_value = df.at[idx, original_name]
                            if pd.notna(original_value) and str(original_value).strip():
                                df.at[idx, duplicated_name] = original_value
                        
                        logger.info(f"‚úÖ [{marketplace}] –°–∏–Ω—Ö—Ä–æ–Ω–∏–∑–∏—Ä–æ–≤–∞–Ω—ã: '{original_name}' ‚Üí '{duplicated_name}'")
            
            # –†–∞—Å—à–∏—Ä—è–µ–º –ª–∏—Å—Ç –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
            current_rows = ws.max_row
            required_rows = data_start_row + len(df)
            
            if required_rows > current_rows:
                print(f"[INFO] –†–∞—Å—à–∏—Ä—è—é –ª–∏—Å—Ç: {current_rows} ‚Üí {required_rows}")
                last_data_row = current_rows
                for row_idx in range(current_rows + 1, required_rows + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        source_cell = ws.cell(row=last_data_row, column=col_idx)
                        target_cell = ws.cell(row=row_idx, column=col_idx)
                        
                        if source_cell.has_style:
                            target_cell.font = source_cell.font.copy()
                            target_cell.border = source_cell.border.copy()
                            target_cell.fill = source_cell.fill.copy()
                            target_cell.number_format = source_cell.number_format
                            target_cell.protection = source_cell.protection.copy()
                            target_cell.alignment = source_cell.alignment.copy()
            
            # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ
            for row_num, (df_row_idx, row) in enumerate(df.iterrows()):
                excel_row_idx = data_start_row + row_num
                
                for df_col_name, value in row.items():
                    if pd.isna(value):
                        continue
                    
                    # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ Excel —Å—Ç–æ–ª–±—Ü–æ–≤ –¥–ª—è —ç—Ç–æ–≥–æ DataFrame —Å—Ç–æ–ª–±—Ü–∞
                    excel_cols = df_to_excel_mapping.get(df_col_name, [])
                    
                    # –ó–∞–ø–∏—Å—ã–≤–∞–µ–º –≤–æ –í–°–ï —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É—é—â–∏–µ —Å—Ç–æ–ª–±—Ü—ã (–¥–ª—è –¥—É–±–ª–∏–∫–∞—Ç–æ–≤)
                    for excel_col_idx in excel_cols:
                        cell = ws.cell(row=excel_row_idx, column=excel_col_idx)
                        
                        # –í–∞–ª–∏–¥–∞—Ü–∏—è —á–µ—Ä–µ–∑ AI (–µ—Å–ª–∏ –µ—Å—Ç—å)
                        allowed_values = self._get_validation_list_values(ws, excel_row_idx, excel_col_idx)
                        
                        if allowed_values and self.ai_comparator:
                            matched_value = self.ai_comparator.match_value_with_list(str(value), allowed_values)
                            if matched_value:
                                cell.value = matched_value
                                stats['ai_matched'] += 1
                            else:
                                stats['validation_conflicts'] += 1
                                stats['skipped'] += 1
                                continue
                        else:
                            cell.value = value
                        
                        stats['saved'] += 1
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∞–π–ª
            wb.save(output_path)
            print(f"[+] {config['display_name']}: —Å–æ—Ö—Ä–∞–Ω–µ–Ω–æ –≤ '{output_path}'")
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        print(f"\n{'='*60}")
        print(f"–°–¢–ê–¢–ò–°–¢–ò–ö–ê:")
        print(f"  ‚úì –ó–∞–ø–∏—Å–∞–Ω–æ: {stats['saved']}")
        if self.ai_comparator:
            print(f"  ü§ñ AI-—Å–æ–ø–æ—Å—Ç–∞–≤–ª–µ–Ω–∏–π: {stats['ai_matched']}")
            print(f"  ‚ö† –ö–æ–Ω—Ñ–ª–∏–∫—Ç–æ–≤: {stats['validation_conflicts']}")
        print(f"{'='*60}")
