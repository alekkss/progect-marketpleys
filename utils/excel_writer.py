"""
Модуль для создания Excel отчетов
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List
from config.config import FILE_CONFIGS
from utils.logger_config import setup_logger
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))


class ExcelWriter:
    """Класс для создания Excel отчетов сравнения столбцов"""
    
    def create_report(self, comparison_result: Dict, output_file: str):
        """
        Создает Excel отчет с результатами сравнения
        
        Args:
            comparison_result: результаты сравнения от AI
            output_file: путь для сохранения файла
        """
        wb = Workbook()
        wb.remove(wb.active)  # Удаляем пустой лист
        
        # Создаем листы
        self._create_summary_sheet(wb, comparison_result)
        self._create_matches_all_three(wb, comparison_result)
        self._create_matches_1_2(wb, comparison_result)
        self._create_matches_1_3(wb, comparison_result)
        self._create_matches_2_3(wb, comparison_result)
        self._create_unique_columns(wb, comparison_result)
        
        # Сохраняем файл
        wb.save(output_file)
        print(f"[+] Отчет сохранен: {output_file}")
    
    def create_report_with_changes(
        self, 
        comparison_result: Dict, 
        changes_log: Dict,
        output_file: str
    ):
        """
        Создает Excel отчет с результатами сравнения и логом изменений
        
        Args:
            comparison_result: результаты сравнения от AI
            changes_log: лог произведенных изменений
            output_file: путь для сохранения файла
        """
        wb = Workbook()
        wb.remove(wb.active)  # Удаляем пустой лист
        
        # Создаем листы со сравнением
        self._create_summary_sheet(wb, comparison_result)
        self._create_matches_all_three(wb, comparison_result)
        self._create_matches_1_2(wb, comparison_result)
        self._create_matches_1_3(wb, comparison_result)
        self._create_matches_2_3(wb, comparison_result)
        self._create_unique_columns(wb, comparison_result)
        
        # НОВОЕ: Создаем листы с логом изменений
        self._create_changes_log_sheets(wb, changes_log)
        
        # Сохраняем файл
        wb.save(output_file)
        print(f"[+] Отчет сохранен с логом изменений: {output_file}")
    
    def _create_changes_log_sheets(self, wb: Workbook, changes_log: Dict):
        """Создает листы с логом изменений для каждого маркетплейса"""
        
        for marketplace, changes in changes_log.items():
            if not changes:
                continue
            
            config = FILE_CONFIGS[marketplace]
            sheet_name = f"Изменения {config['display_name']}"
            
            # Ограничиваем длину названия листа (max 31 символ)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + "..."
            
            ws = wb.create_sheet(sheet_name)
            
            # Заголовки
            headers = ['Артикул', 'Столбец', 'Новое значение']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Данные
            for row_num, change in enumerate(changes, 2):
                article = change.get('article', '')
                column = change.get('column', '')
                new_value = change.get('new_value', '')
                
                # Артикул
                cell = ws.cell(row=row_num, column=1, value=article)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Столбец
                cell = ws.cell(row=row_num, column=2, value=column)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Новое значение
                cell = ws.cell(row=row_num, column=3, value=new_value)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Чередующиеся цвета строк для лучшей читаемости
                if row_num % 2 == 0:
                    fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    for col in [1, 2, 3]:
                        ws.cell(row=row_num, column=col).fill = fill
            
            # Ширина столбцов
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 40
            
            # Замораживаем первую строку
            ws.freeze_panes = 'A2'
            
            # Статистика
            print(f"[+] Лист '{sheet_name}': записано {len(changes)} изменений")
    
    def _create_summary_sheet(self, wb: Workbook, comparison_result: Dict):
        """Создает лист с общей статистикой"""
        ws = wb.create_sheet("Сводка", 0)
        
        # Заголовок
        ws['A1'] = "СВОДКА СРАВНЕНИЯ МАРКЕТПЛЕЙСОВ"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        # Статистика
        stats = [
            ("Совпадения во всех 3 маркетплейсах", len(comparison_result.get('matches_all_three', []))),
            ("Совпадения WB ↔ Ozon", len(comparison_result.get('matches_1_2', []))),
            ("Совпадения WB ↔ Яндекс", len(comparison_result.get('matches_1_3', []))),
            ("Совпадения Ozon ↔ Яндекс", len(comparison_result.get('matches_2_3', []))),
            ("Уникальные столбцы в WB", len(comparison_result.get('only_in_first', []))),
            ("Уникальные столбцы в Ozon", len(comparison_result.get('only_in_second', []))),
            ("Уникальные столбцы в Яндекс", len(comparison_result.get('only_in_third', []))),
        ]
        
        row = 3
        for label, count in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count
            ws[f'B{row}'].font = Font(bold=True, size=12)
            ws.row_dimensions[row].height = 25
            row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
    
    def _create_matches_all_three(self, wb: Workbook, comparison_result: Dict):
        """Создает лист с совпадениями всех трех маркетплейсов"""
        ws = wb.create_sheet("Совпадения (все 3)")
        
        # Заголовки
        headers = ['WB', 'Ozon', 'Яндекс', 'Уверенность']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        
        # Данные
        for row_num, match in enumerate(comparison_result.get('matches_all_three', []), 2):
            ws.cell(row=row_num, column=1, value=match.get('column_1', ''))
            ws.cell(row=row_num, column=2, value=match.get('column_2', ''))
            ws.cell(row=row_num, column=3, value=match.get('column_3', ''))
            
            confidence = int(match.get('confidence', 0) * 100)
            ws.cell(row=row_num, column=4, value=f"{confidence}%")
        
        self._format_sheet(ws, len(comparison_result.get('matches_all_three', [])))
    
    def _create_matches_1_2(self, wb: Workbook, comparison_result: Dict):
        """Создает лист с совпадениями WB ↔ Ozon"""
        ws = wb.create_sheet("WB ↔ Ozon")
        
        # Заголовки
        headers = ['WB', 'Ozon', 'Уверенность']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Данные
        for row_num, match in enumerate(comparison_result.get('matches_1_2', []), 2):
            ws.cell(row=row_num, column=1, value=match.get('column_1', ''))
            ws.cell(row=row_num, column=2, value=match.get('column_2', ''))
            
            confidence = int(match.get('confidence', 0) * 100)
            ws.cell(row=row_num, column=3, value=f"{confidence}%")
        
        self._format_sheet(ws, len(comparison_result.get('matches_1_2', [])))
    
    def _create_matches_1_3(self, wb: Workbook, comparison_result: Dict):
        """Создает лист с совпадениями WB ↔ Яндекс"""
        ws = wb.create_sheet("WB ↔ Яндекс")
        
        # Заголовки
        headers = ['WB', 'Яндекс', 'Уверенность']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Данные
        for row_num, match in enumerate(comparison_result.get('matches_1_3', []), 2):
            ws.cell(row=row_num, column=1, value=match.get('column_1', ''))
            ws.cell(row=row_num, column=2, value=match.get('column_3', ''))
            
            confidence = int(match.get('confidence', 0) * 100)
            ws.cell(row=row_num, column=3, value=f"{confidence}%")
        
        self._format_sheet(ws, len(comparison_result.get('matches_1_3', [])))
    
    def _create_matches_2_3(self, wb: Workbook, comparison_result: Dict):
        """Создает лист с совпадениями Ozon ↔ Яндекс"""
        ws = wb.create_sheet("Ozon ↔ Яндекс")
        
        # Заголовки
        headers = ['Ozon', 'Яндекс', 'Уверенность']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        
        # Данные
        for row_num, match in enumerate(comparison_result.get('matches_2_3', []), 2):
            ws.cell(row=row_num, column=1, value=match.get('column_2', ''))
            ws.cell(row=row_num, column=2, value=match.get('column_3', ''))
            
            confidence = int(match.get('confidence', 0) * 100)
            ws.cell(row=row_num, column=3, value=f"{confidence}%")
        
        self._format_sheet(ws, len(comparison_result.get('matches_2_3', [])))
    
    def _create_unique_columns(self, wb: Workbook, comparison_result: Dict):
        """Создает лист с уникальными столбцами"""
        ws = wb.create_sheet("Уникальные")
        
        # Заголовки
        headers = ['WB', 'Ozon', 'Яндекс']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        # Данные
        max_rows = max(
            len(comparison_result.get('only_in_first', [])),
            len(comparison_result.get('only_in_second', [])),
            len(comparison_result.get('only_in_third', []))
        )
        
        for row_num in range(max_rows):
            wb_cols = comparison_result.get('only_in_first', [])
            ozon_cols = comparison_result.get('only_in_second', [])
            yandex_cols = comparison_result.get('only_in_third', [])
            
            if row_num < len(wb_cols):
                ws.cell(row=row_num + 2, column=1, value=wb_cols[row_num])
            if row_num < len(ozon_cols):
                ws.cell(row=row_num + 2, column=2, value=ozon_cols[row_num])
            if row_num < len(yandex_cols):
                ws.cell(row=row_num + 2, column=3, value=yandex_cols[row_num])
        
        self._format_sheet(ws, max_rows)
    
    def _format_sheet(self, ws, data_rows: int):
        """Форматирует лист"""
        # Ширина столбцов
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 35
        
        # Замораживаем первую строку
        ws.freeze_panes = 'A2'
        
        # Чередующиеся цвета строк
        for row_num in range(2, data_rows + 2):
            if row_num % 2 == 0:
                fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                for cell in ws[row_num]:
                    if cell.value is not None:
                        cell.fill = fill
